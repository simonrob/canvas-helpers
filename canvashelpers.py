"""Utility functions for Canvas helper scripts."""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2024 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2024-04-05'  # ISO 8601 (YYYY-MM-DD)

import configparser
import csv
import json
import os
import re
import sys
import tempfile

import openpyxl
import requests.structures


class Config:
    FILE_PATH = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'canvashelpers.config')

    configparser = configparser.ConfigParser()
    configparser.read(FILE_PATH)
    SETTINGS = configparser[configparser.sections()[0]]

    API_TOKEN = SETTINGS['canvas_api_token']  # all scripts need this token; only a subset need the full settings below

    if API_TOKEN.startswith('*** your'):
        print('WARNING: API token in', FILE_PATH, 'seems to contain the example value - please make sure you have',
              'added your own token')

    @staticmethod
    def get_settings():
        return Config.SETTINGS


class Utils:
    @staticmethod
    def course_url_to_api(url):
        return url.rstrip('/').replace('/courses', '/api/v1/courses')

    @staticmethod
    def course_url_to_speedgrader(url, add_student_id=None):
        speedgrader_url = url.replace('assignments/', 'gradebook/speed_grader?assignment_id=')
        if add_student_id:
            speedgrader_url += '&student_id=' + str(add_student_id)
        return speedgrader_url

    @staticmethod
    def get_course_id(course_url):
        return int(course_url.split('courses/')[-1].split('/')[0])

    @staticmethod
    def get_assignment_id(assignment_url):
        return int(assignment_url.rstrip('/').split('/')[-1])

    @staticmethod
    def get_user_details(api_root, user_id='self'):
        user_details_response = requests.get('%s/users/%s/' % (api_root, user_id), headers=Utils.canvas_api_headers())
        if user_details_response.status_code != 200:
            return user_id, 'UNKNOWN NAME'
        user_details_json = user_details_response.json()
        return user_details_json['id'], user_details_json['name']

    @staticmethod
    def ordered_strings(text):
        # used to sort a list of numbers and/or names in a more natural order
        return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', text)]

    @staticmethod
    def canvas_api_headers():
        submission_list_headers = requests.structures.CaseInsensitiveDict()
        submission_list_headers['accept'] = 'application/json'
        submission_list_headers['authorization'] = 'Bearer %s' % Config.API_TOKEN
        return submission_list_headers

    @staticmethod
    def canvas_multi_page_request(current_request_url, params=None, type_hint='API'):
        """Retrieve a full (potentially multi-page) response from the Canvas API. If the initial response refers to
        subsequent pages of results, these are loaded and concatenated automatically. For (slightly) more specific
        progress/error messages, set type_hint to a string describing the API call that is being made """
        if not params:
            params = {}
        params['per_page'] = 100
        response = '[]'
        while True:
            print('Requesting', type_hint, 'page:', current_request_url)
            current_response = requests.get(current_request_url, params=params, headers=Utils.canvas_api_headers())
            if current_response.status_code != 200:
                print('ERROR: unable to load complete', type_hint, 'response - status code',
                      current_response.status_code)
                return None

            response = response[:-1] + ',' + current_response.text[1:]

            # see: https://canvas.instructure.com/doc/api/file.pagination.html
            page_links = current_response.headers['Link'] if 'Link' in current_response.headers else ''
            next_page_match = re.search(r',\s*<(?P<next>.*?)>;\s*rel="next"', page_links)
            if next_page_match:
                current_request_url = next_page_match.group('next')
            else:
                return '[' + response[2:]

    @staticmethod
    def get_course_users(course_url, includes=None, enrolment_types=None):
        """Get a list of users in a course, returning a string that can be parsed as JSON. This function is simply
        a wrapper around Utils.canvas_multi_page_request, but is kept to separate the API parameter complexity from
        the scripts that use this method"""
        params = {'enrollment_type[]': ['student'] if not enrolment_types else enrolment_types}
        if includes:
            params['include[]'] = []
            for param in includes:
                params['include[]'].append(param)
        return Utils.canvas_multi_page_request('%s/users' % course_url, params=params,
                                               type_hint='course users list')

    @staticmethod
    def get_course_enrolments(course_url, includes=None, enrolment_types=None):
        """Get a list of enrolments in a course, which is slightly different to get_course_users in that it allows us to
        identify the inbuilt test student via their enrolment type 'StudentViewEnrollment'. This function is simply
        a wrapper around Utils.canvas_multi_page_request, but is kept to separate the API parameter complexity from
        the scripts that use this method"""
        params = {'type[]': ['StudentViewEnrollment'] if not enrolment_types else enrolment_types}
        if includes:
            params['include[]'] = []
            for param in includes:
                params['include[]'].append(param)
        return Utils.canvas_multi_page_request('%s/enrollments' % course_url, params=params,
                                               type_hint='filtered course enrolments list')

    @staticmethod
    def get_course_groups(course_group_tab_url, group_by='group_number'):
        # noinspection GrazieInspection
        """Get details of all groups within a group set. Pass the URL of the desired group set as shown in the web
        interface (i.e., https://canvas.instructure.com/courses/[course ID]/groups#tab-[group ID]). Note that groups
        *must* be named in the format [name][space][number] (i.e., "Group 1", "Group 2", etc). This API endpoint is
        currently a beta method, and not always reliable, so we also include an iteration approach. Returns a tuple of
        (group set ID, group set dict). The `group_by` parameter can either be `group_number` (default) for the integer
        sequence number of each group; `group_name` for the full name string; or, any other value for student numbers"""
        group_set_id = course_group_tab_url.split('#tab-')[-1]
        group_sets = {}
        try:
            group_set_id = int(group_set_id)
        except ValueError:
            print('ERROR: unable to get group set ID from given URL', course_group_tab_url)
            return None, None

        csv_headers = None
        api_url = Utils.course_url_to_api(course_group_tab_url).split('/courses')[0]
        group_set_response = requests.get('%s/group_categories/%d/export' % (api_url, group_set_id),
                                          headers=Utils.canvas_api_headers())
        if group_set_response.status_code != 200:
            if group_set_response.status_code == 401:
                # archived courses don't support this method, so we use the old iterative approach
                print('WARNING: unable to bulk export group set data; switching to legacy iteration method')
                return Utils._get_course_groups_legacy(course_group_tab_url, group_by)
            else:
                print('ERROR: unable to load group set', group_set_id, '- aborting',
                      '(error:', group_set_response.text, ')')
                sys.exit()

        group_cache_file = tempfile.NamedTemporaryFile(mode='w', delete=False)
        cache_file_name = group_cache_file.name
        group_cache_file.write(group_set_response.text)
        group_cache_file.close()
        with open(cache_file_name) as group_cache_file:
            reader = csv.reader(group_cache_file)
            for row in reader:
                if not csv_headers:
                    csv_headers = row
                    continue

                try:
                    # skip non-students, often with non-numeric IDs (but intentionally keep as a string) for later use
                    # as Canvas is inconsistent with its treatment of these (e.g., course users: int; groups: string)
                    int(row[csv_headers.index('login_id')])
                except ValueError:
                    print('WARNING: skipping non-numeric group member login_id:', row[csv_headers.index('login_id')])
                    continue
                if not row[csv_headers.index('group_name')]:  # course members not in a group have an empty group name
                    print('WARNING: skipping course member not in any group:', row[csv_headers.index('login_id')])
                    continue

                group_entry = {
                    'group_name': row[csv_headers.index('group_name')],
                    'group_id': row[csv_headers.index('canvas_group_id')],
                    'group_number': int(row[csv_headers.index('group_name')].split(' ')[-1]),
                    'student_number': row[csv_headers.index('login_id')],
                    'student_name': row[csv_headers.index('name')],
                    'student_canvas_id': row[csv_headers.index('canvas_user_id')]
                }

                if group_by in ['group_number', 'group_name']:
                    if group_entry[group_by] not in group_sets:
                        group_sets[group_entry[group_by]] = []
                    group_sets[group_entry[group_by]].append(group_entry)
                else:
                    group_sets[group_entry['student_number']] = group_entry
        os.remove(cache_file_name)

        print('Loaded', len(group_sets), 'valid group records from', course_group_tab_url)
        return group_set_id, dict(sorted(group_sets.items()))

    @staticmethod
    def _get_course_groups_legacy(course_group_tab_url, group_by):
        group_set_id = course_group_tab_url.split('#tab-')[-1]
        group_sets = {}
        try:
            group_set_id = int(group_set_id)
        except ValueError:
            print('ERROR: unable to get group set ID from given URL', course_group_tab_url)
            return None, None

        api_url = Utils.course_url_to_api(course_group_tab_url).split('/courses')[0]
        group_set_response = Utils.canvas_multi_page_request('%s/group_categories/%d/groups' % (api_url, group_set_id),
                                                             type_hint='group sets')
        if not group_set_response:
            print('ERROR: unable to load group sets; aborting')
            sys.exit()

        group_set_json = json.loads(group_set_response)
        for group in group_set_json:
            group_members_response = Utils.canvas_multi_page_request('%s/groups/%d/users' % (api_url, group['id']),
                                                                     type_hint='group')
            if not group_members_response:
                print('WARNING: unable to load group members; skipping group', group)
                continue

            group_members_json = json.loads(group_members_response)
            for member in group_members_json:
                try:
                    int(member['login_id'])  # ignore non-students, who often have non-numeric IDs
                except ValueError:
                    print('WARNING: skipping non-numeric group member login_id:', member['login_id'])
                    continue

                group_entry = {
                    'group_name': group['name'],
                    'group_id': int(group['id']),
                    'group_number': int(group['name'].split(' ')[-1]),
                    'student_number': int(member['login_id']),
                    'student_name': member['name'],
                    'student_canvas_id': int(member['id'])
                }

                if group_by in ['group_number', 'group_name']:
                    if group_entry[group_by] not in group_sets:
                        group_sets[group_entry[group_by]] = []
                    group_sets[group_entry[group_by]].append(group_entry)
                else:
                    group_sets[group_entry['student_number']] = group_entry

        print('Loaded', len(group_sets), 'valid group records from', course_group_tab_url)
        return group_set_id, dict(sorted(group_sets.items()))

    @staticmethod
    def get_assignment_submissions(assignment_url, includes=None):
        """Get a list of assignment submissions, returning a string that can be parsed as JSON. This function is simply
        a wrapper around Utils.canvas_multi_page_request, but is kept to separate the API parameter complexity from
        the scripts that use this method"""
        # TODO: handle variants (include[]=submission_history): canvas.instructure.com/doc/api/submissions.html
        # TODO: does requesting group option when there are no groups cause any problems? (no issues seen so far)
        # see: https://canvas.instructure.com/doc/api/submissions.html#method.submissions_api.index
        params = {'include[]': []}
        includes = ['user', 'group'] + (includes if includes else [])
        for param in includes:
            params['include[]'].append(param)
        return Utils.canvas_multi_page_request('%s/submissions' % assignment_url, params=params,
                                               type_hint='assignment submissions list')

    @staticmethod
    def filter_assignment_submissions(assignment_url, submission_list_json, groups_mode=False,
                                      include_unsubmitted=False, ignored_users=None, sort_entries=False):
        """Filter a list of submissions (in parsed JSON format). Setting groups_mode to True will remove any users who
        are not in a group, and skip any duplicates (which occur because Canvas associates group submissions with each
        group member individually). Setting include_unsubmitted to True will include all entries, even those that do
        not actually have a submission. The ignored_users parameter is an array of Canvas user IDs, and is used to
        remove specific submitters (typically the inbuilt test users)"""
        filtered_submission_list = []
        for submission in submission_list_json:
            ignored_submission = False
            # TODO: sometimes groups without submissions do not appear at all in the submission list - is this fixable?
            if ('workflow_state' in submission and submission['workflow_state'] == 'unsubmitted') \
                    or 'workflow_state' not in submission:
                if not include_unsubmitted:
                    ignored_submission = True

            if groups_mode and not ignored_submission:
                if submission['group']['id'] is None:
                    ignored_submission = True
                else:
                    for parsed_submission in filtered_submission_list:
                        if submission['group']['id'] == parsed_submission['group']['id']:
                            ignored_submission = True
                            break

            if ignored_users and submission['user_id'] in ignored_users:
                ignored_submission = True

            if not ignored_submission:
                if 'login_id' not in submission['user']:
                    # this is the only reason to have the assignment URL in this function
                    submission['user']['login_id'] = Utils.get_canvas_user_login_id(assignment_url,
                                                                                    submission['user']['id'])
                filtered_submission_list.append(submission)

        if sort_entries:
            filtered_submission_list = sorted(filtered_submission_list,
                                              key=lambda entry: Utils.ordered_strings(
                                                  entry['group']['name'] if groups_mode else entry['user']['login_id']))

        print('Loaded', 'and sorted' if sort_entries else '', len(filtered_submission_list), 'valid submissions',
              '(discarded', (len(submission_list_json) - len(filtered_submission_list)),
              'filtered, duplicate, invalid/incomplete or missing)')
        return filtered_submission_list

    @staticmethod
    def get_submitter_details(assignment_url, submission, groups_mode=False):
        """For a given submission object (in parsed JSON format), return the submitter's details (the Canvas ID of
        the user who submitted, their Login ID (typically institutional student number), and their name). Setting
        groups_mode to True will return the Canvas group ID and the group name instead of Login ID and student name.
        There is currently no handling of users who are not part of a group (whose group attributes will be None);
        however, if Utils.filter_assignment_submissions is used (with groups_mode=True) beforehand then these users
        will not be present regardless"""
        submitter = None
        if groups_mode:
            if 'group' in submission:
                submitter = {'canvas_user_id': submission['user_id'], 'canvas_group_id': submission['group']['id'],
                             'group_name': submission['group']['name']}
                if 'login_id' in submission['user']:
                    # login_id is not always present (perhaps linked to individual marks in group assignments)
                    submitter['student_number'] = submission['user']['login_id']
                else:
                    # this is the only reason to have the assignment URL in this function
                    submitter['student_number'] = Utils.get_canvas_user_login_id(assignment_url,
                                                                                 submission['user']['id'])
        elif 'user' in submission:
            submitter = {'canvas_user_id': submission['user_id'], 'student_number': submission['user']['login_id'],
                         'student_name': submission['user']['name']}
        return submitter

    @staticmethod
    def get_marks_mapping(marks_file):
        marks_map = {}
        if os.path.exists(marks_file):
            if marks_file.lower().endswith('.xlsx'):
                marks_workbook = openpyxl.load_workbook(marks_file)
                marks_sheet = marks_workbook[marks_workbook.sheetnames[0]]
                for row in marks_sheet.iter_rows():
                    Utils.parse_marks_file_row(marks_map, [entry.value for entry in row])
            else:
                with open(marks_file, newline='') as marks_csv:
                    reader = csv.reader(marks_csv)
                    for row in reader:
                        Utils.parse_marks_file_row(marks_map, row)
        return marks_map

    @staticmethod
    def get_assignment_student_list(assignment_url):
        """For a given assignment, get the list of students it is assigned to. In most cases it is better to use
        Utils.get_assignment_submissions, which returns users as part of its main response. However, the New Quizzes
        API does not return Login IDs, so for that script this method is used to match submissions instead"""
        params = {'include[]': ['enrollments']}
        user_list_response = Utils.canvas_multi_page_request('%s/users' % assignment_url.split('/assignments')[0],
                                                             params=params, type_hint='assignment student list')
        if not user_list_response:
            return None

        user_list_json = json.loads(user_list_response)
        submission_student_map = []
        for user in user_list_json:
            for role in user['enrollments']:
                if role['type'] == 'StudentEnrollment' and role['enrollment_state'] == 'active':
                    if 'login_id' in user:
                        student_number = user['login_id']
                    else:
                        student_number = Utils.get_canvas_user_login_id(assignment_url, user['id'])
                    submission_student_map.append({'student_number': student_number, 'user_id': user['id']})
        return submission_student_map

    @staticmethod
    def get_canvas_user_login_id(assignment_url, user_id):
        # Canvas has a bug where login_id is missing in some requests - need to get individually (slowly...)
        print('WARNING: encountered Canvas bug in user list; requesting profile for', user_id, 'individually')
        user_profile_response = requests.get('%s/users/%s/profile' % (assignment_url.split('/courses')[0], user_id),
                                             headers=Utils.canvas_api_headers())
        if user_profile_response.status_code != 200:
            print('ERROR: unable to load user profile for', user_id)
            return None  # TODO: is there anything else we can do?
        else:
            return user_profile_response.json()['login_id']

    @staticmethod
    def parse_marks_file_row(marks_map, row):
        # ultra-simplistic check to avoid any header rows (headers are not normally numeric)
        try:
            grade = float(row[1])
        except (ValueError, TypeError):
            return

        student_number_or_group_name = str(row[0])
        marks_map_entry = {'mark': grade}
        if len(row) > 2 and row[2]:  # individual comment is optional
            marks_map_entry['comment'] = row[2]

        if student_number_or_group_name:
            marks_map[student_number_or_group_name] = marks_map_entry


class Args:
    @staticmethod
    def interactive(f):
        """This rather complicated setup is to allow usage of both Tooey and Gooey at the same time when it is possible
        that one, both or neither are available to import"""
        tooey_ignore = '--ignore-tooey'
        gooey_ignore = '--ignore-gooey'
        has_gooey = False
        using_gooey = False
        try:
            # noinspection PyPackageRequirements,PyUnresolvedReferences
            import gooey
            has_gooey = True
            if gooey_ignore not in sys.argv:
                using_gooey = True
                return gooey.Gooey(f)()
        except ImportError:
            pass

        if not using_gooey:
            try:
                # noinspection PyPackageRequirements,PyUnresolvedReferences
                import tooey
                if tooey_ignore not in sys.argv:
                    if gooey_ignore in sys.argv:
                        sys.argv.remove(gooey_ignore)
                    return tooey.Tooey(f)()
            except ImportError:
                pass

        if tooey_ignore in sys.argv:
            sys.argv.remove(tooey_ignore)

        if has_gooey:
            # a successful run with Gooey calls the program again with the actual arguments plus `--ignore-gooey`
            # noinspection PyUnboundLocalVariable
            return gooey.Gooey(f)()
        else:
            if gooey_ignore in sys.argv:
                sys.argv.remove(gooey_ignore)
            return f()
