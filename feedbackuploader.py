"""It can often be useful to mark assignments and produce feedback text or attachments offline, rather than directly
in Canvas. When doing so using the SpeedGrader it is very time-consuming to add these items one-by-one. This script
lets you upload a set of attachments, feedback comments and marks in bulk."""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2023 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2023-06-05'  # ISO 8601 (YYYY-MM-DD)

import argparse
import csv
import json
import mimetypes
import os
import sys

import openpyxl
import requests

from canvashelpers import Config, Utils

parser = argparse.ArgumentParser()
parser.add_argument('url', nargs=1,
                    help='Please pass the URL of the assignment to bulk upload feedback attachments for')
parser.add_argument('--working-directory', default=None,
                    help='The root directory to use for the script\'s operation. Within this directory, attachments '
                         'and any `--marks-file` should be placed in a subfolder named as the assignment number (e.g., '
                         'for an assignment at https://[canvas-domain]/courses/10000/assignments/123456, name the '
                         'subfolder 123456). Default: the same directory as this script')
parser.add_argument('--attachment-extension', default='pdf',
                    help='The file extension of the attachments to upload (without the dot separator). Attachments '
                         'should be named following the format [student number].[extension] (or, in group mode, '
                         '[group name].[extension]. Multiple attachments can be added by running the script '
                         'repeatedly. Default: \'pdf\'')
parser.add_argument('--attachment-mime-type', default=None,
                    help='Canvas requires a hint about the MIME type of the attachment file you are uploading. The '
                         'script is able to guess the correct value in most cases, but if you are uploading a file '
                         'with an unusual extension or format then you can specify a value here')
parser.add_argument('--marks-file', default=None,
                    help='An XLSX or CSV file containing a minimum of two columns: student number and mark (in that '
                         'order). A third column can be added for per-student feedback that will be added as a text '
                         'comment, overriding the global `--attachment-comment`. To add a comment but not a mark, set '
                         'a negative mark (e.g., -1). Please note that unless `--marks-as-percentage` is set, the mark '
                         'must be on the same scale as that of the assignment itself. The script tries to validate '
                         'this, but is not always able to do so if marks are low')
parser.add_argument('--marks-as-percentage', action='store_true',
                    help='Set this parameter if your `--marks-file` marks are provided as a percentage, rather than on '
                         'the same scale as the marks available for the assignment itself')
parser.add_argument('--attachment-comment', default='See attached file',
                    help='The comment to add when attaching the feedback file. Overridden by any individual comments '
                         'in the imported marks file. The default value (\'See attached file\') will be skipped if '
                         'there is no attachment, but in all other cases the comment will be added regardless')
parser.add_argument('--groups', action='store_true',
                    help='Use this option if the assignment is completed in groups and all members should receive the '
                         'same mark and feedback. If you use this option, group names must be used instead of student '
                         'numbers in both the feedback filenames and any `--marks-file` attachment, and these must be '
                         'exactly as specified on Canvas. For example, if you have a Canvas group called \'Group 1\', '
                         'name the attachment file \'Group 1.pdf\'; (\'1.pdf\' will not work)')
parser.add_argument('--groups-individual', action='store_true',
                    help='Use this option *in addition* to `--groups` if the assignment is completed in groups but it '
                         'is configured to give group members marks and feedback individually. The script will first '
                         'look for attachments or `--marks-file` entries named after the group (for cases where all '
                         'members should receive the same feedback). If these items are not found, the script will '
                         'then look for attachments or `--marks-file` entries named after student numbers (for cases '
                         'when individual feedback is needed). These approaches can be mixed if needed (e.g., a '
                         'group-named attachment but individual marks and/or feedback comments).')
parser.add_argument('--include-unsubmitted', action='store_true',
                    help='Students who have not made a submission for the assignment are skipped by default. Set this '
                         'option if you want to include these students (for example, when no submission is actually '
                         'expected, and the Canvas assignment is used solely to record marks). Note that when not in '
                         '`--groups` mode this will include any staff enrolled as students (though not the inbuilt '
                         'test student), but this should not be an issue as no mark, comment or attachment will be '
                         'available for them')
parser.add_argument('--delete-existing', action='store_true',
                    help='Delete all existing comments created by your Canvas user before adding any new feedback '
                         '(removing both manually-created comments and ones added via API scripts such as this one). '
                         'If comments have attachments, the attachments will also become inaccessible. Note that this '
                         'option does not change any marks that have been entered; only comments are removed.')
parser.add_argument('--dry-run', action='store_true',
                    help='Preview the script\'s actions without actually making any changes. Highly recommended!')
args = parser.parse_args()  # exits if no assignment URL is provided

ASSIGNMENT_URL = Utils.course_url_to_api(args.url[0])
assignment_id = Utils.get_assignment_id(ASSIGNMENT_URL)
INPUT_DIRECTORY = os.path.join(
    os.path.dirname(os.path.realpath(__file__)) if args.working_directory is None else args.working_directory,
    str(assignment_id))
if not os.path.exists(INPUT_DIRECTORY):
    print('ERROR: input directory not found - please place all files to upload (and any `--marks-file`) in the '
          'folder %s' % INPUT_DIRECTORY)
    sys.exit()
print('%sUploading assignment feedback from %s to assignment %s' % (
    'DRY RUN: ' if args.dry_run else '', INPUT_DIRECTORY, args.url[0]))

marks_map = {}
if args.marks_file:
    marks_file = os.path.join(INPUT_DIRECTORY, args.marks_file)
    if os.path.exists(marks_file):
        if marks_file.lower().endswith('.xlsx'):
            marks_workbook = openpyxl.load_workbook(marks_file)
            marks_sheet = marks_workbook[marks_workbook.sheetnames[0]]
            for row in marks_sheet.iter_rows():
                Utils.parse_marks_file_row(marks_map, [entry.value for entry in row])
        else:
            with open(marks_file, newline='') as marks_csv:
                reader = csv.reader(marks_csv)
                for row in reader:
                    Utils.parse_marks_file_row(marks_map, row)
        print('Loaded marks/feedback mapping for', len(marks_map), 'submissions:', marks_map)
    else:
        print('Ignoring marks file argument', args.marks_file, '- not found in assignment directory at', marks_file)

assignment_details_response = requests.get(ASSIGNMENT_URL, headers=Utils.canvas_api_headers())
if assignment_details_response.status_code != 200:
    print('ERROR: unable to get assignment details - did you set a valid Canvas API token in %s?' % Config.FILE_PATH)
    sys.exit()
maximum_marks = assignment_details_response.json()['points_possible']
mark_exceeded = False
for mark_row in marks_map:
    if marks_map[mark_row]['mark'] > maximum_marks and not args.marks_as_percentage:
        print('ERROR: marks file entry for', mark_row, 'awards more than the maximum', maximum_marks, 'marks available',
              '-', marks_map[mark_row])
        mark_exceeded = True
if mark_exceeded:
    sys.exit()

submission_list_response = Utils.get_assignment_submissions(ASSIGNMENT_URL, includes=['submission_comments'])
if not submission_list_response:
    print('ERROR: unable to retrieve submission list; aborting')
    sys.exit()

# identify and ignore the inbuilt test student
course_enrolment_response = Utils.get_course_enrolments(ASSIGNMENT_URL.split('/assignments')[0])
if not course_enrolment_response:
    print('ERROR: unable to retrieve course enrolment list; aborting')
    sys.exit()
ignored_users = [user['user_id'] for user in json.loads(course_enrolment_response)]

submission_list_json = json.loads(submission_list_response)
filtered_submission_list = Utils.filter_assignment_submissions(submission_list_json,
                                                               groups_mode=args.groups and not args.groups_individual,
                                                               include_unsubmitted=args.include_unsubmitted,
                                                               ignored_users=ignored_users, sort_entries=True)

if args.delete_existing:
    print('\nDeleting existing submission comments created by your Canvas user')
    SELF_ID, user_name = Utils.get_user_details(ASSIGNMENT_URL.split('/courses')[0], user_id='self')
    if not SELF_ID:
        print('\tERROR: unable to retrieve your Canvas ID; aborting')
        sys.exit()

    skipped_comments = 0
    for submission in filtered_submission_list:
        if 'submission_comments' in submission:
            for comment in submission['submission_comments']:
                if comment['author_id'] != SELF_ID:
                    skipped_comments += 1
                    continue

                if args.dry_run:
                    print('\tDRY RUN: skipping deletion of existing comment:', comment)
                    continue

                comment_deletion_url = '%s/submissions/%d/comments/%d' % (
                    ASSIGNMENT_URL, submission['user_id'], comment['id'])
                comment_deletion_response = requests.delete(comment_deletion_url, headers=Utils.canvas_api_headers())
                if comment_deletion_response.status_code == 200:
                    print('\tDeleted existing submission comment:', comment)
                else:
                    print('\tWARNING: unable to delete existing submission comment:', comment_deletion_response.text)

    if skipped_comments > 0:
        print('\tSkipped deletion of', skipped_comments, 'existing comments created by other users')

submission_count = 0
submission_total = len(filtered_submission_list)
for submission in filtered_submission_list:
    submission_count += 1
    submitter = Utils.get_submitter_details(submission, groups_mode=args.groups)
    if not submitter:
        print('WARNING: submitter details not found for submission; skipping:', submission)
        continue

    print('\nProcessing submission', submission_count, 'of', submission_total, 'from', submitter)
    user_submission_url = '%s/submissions/%d' % (ASSIGNMENT_URL, submitter['canvas_user_id'])

    feedback_identifier = submitter['group_name'] if args.groups else submitter['student_number']
    attachment_file = '%s.%s' % (feedback_identifier, args.attachment_extension)
    attachment_path = os.path.join(INPUT_DIRECTORY, attachment_file)
    attachment_mime_type = args.attachment_mime_type or mimetypes.guess_type(attachment_path)[0]
    attachment_exists = os.path.exists(attachment_path)

    if attachment_exists and attachment_mime_type:
        print('Found submission attachment file', attachment_file, 'with MIME type', attachment_mime_type)
    elif args.groups and args.groups_individual:  # groups mode but with potential for individual feedback attachment
        attachment_file = '%s.%s' % (submitter['student_number'], args.attachment_extension)
        attachment_path = os.path.join(INPUT_DIRECTORY, attachment_file)
        attachment_mime_type = args.attachment_mime_type or mimetypes.guess_type(attachment_path)[0]
        attachment_exists = os.path.exists(attachment_path)

        if attachment_exists and attachment_mime_type:
            print('Found individual group member submission attachment file', attachment_file, 'with MIME type',
                  attachment_mime_type)
        else:
            print('Both group (%s.%s)' % (submitter['group_name'], args.attachment_extension),
                  'and individual (%s)' % attachment_file, 'attachment at %s' % os.path.dirname(attachment_path),
                  'were not found or are not of a recognised MIME type; skipping upload for this submission')
            attachment_file = None
    else:
        print('Attachment %s at %s' % (attachment_file, os.path.dirname(attachment_path)),
              'not found;' if not attachment_exists else 'is not of a recognised MIME type;',
              'skipping upload for this submission')
        attachment_file = None

    # filter out unset fields, allowing any combination of mark/comment/attachment)
    attachment_comment = args.attachment_comment
    attachment_mark = None
    if feedback_identifier not in marks_map and args.groups and args.groups_individual:
        # groups mode but with potential for individual feedback if no group feedback was found
        feedback_identifier = submitter['student_number']
    if feedback_identifier in marks_map:
        marks_map[feedback_identifier]['matched'] = True
        attachment_mark = marks_map[feedback_identifier]['mark']
        if attachment_mark < 0:
            attachment_mark = None
            print('Spreadsheet mark is < 0; skipping posting a mark for this submission')
        if 'comment' in marks_map[feedback_identifier]:
            attachment_comment = marks_map[feedback_identifier]['comment']
    elif attachment_file is None:
        print('Could not find attachment, mark or comment for submission (at least one item is required); skipping')
        continue
    else:
        print('No entry found in mark/comment spreadsheet for', feedback_identifier)

    # see: https://canvas.instructure.com/doc/api/submissions.html#method.submissions_api.update
    comment_association_data = {'comment[text_comment]': attachment_comment.replace('\\n', '\n')}
    if args.groups and not args.groups_individual:
        comment_association_data['comment[group_comment]'] = True
    if attachment_comment != args.attachment_comment:
        print('Adding submission comment from spreadsheet:', attachment_comment.replace('\n', '\\n'))
    else:
        if attachment_file is None and attachment_comment == parser.get_default('attachment_comment'):
            print('Skipping default comment \'%s\' as no attachment is provided' % attachment_comment)
            del comment_association_data['comment[text_comment]']
        else:
            print('Using attachment comment provided as script argument:', attachment_comment)

    if attachment_mark is not None:
        comment_association_data['submission[posted_grade]'] = attachment_mark
        if args.marks_as_percentage:
            comment_association_data['submission[posted_grade]'] = '%s%%' % attachment_mark
        print('Adding submission mark from spreadsheet:', comment_association_data['submission[posted_grade]'])

    if args.dry_run:
        print('DRY RUN: skipping attachment upload and comment posting steps; moving to next submission')
        continue

    if attachment_file:
        # if there is an attachment we first need to request an upload URL, then associate with a submission comment
        submission_form_data = {'name': attachment_file, 'content_type': attachment_mime_type}
        file_submission_url_response = requests.post('%s/comments/files' % user_submission_url,
                                                     data=submission_form_data, headers=Utils.canvas_api_headers())
        if file_submission_url_response.status_code != 200:
            print('\tERROR: unable to retrieve attachment upload URL; skipping submission')
            continue

        file_submission_url_json = file_submission_url_response.json()
        print('\tUploading feedback attachment to', file_submission_url_json['upload_url'].split('?')[0], '[truncated]')

        files_data = {'file': (attachment_file, open(attachment_path, 'rb'))}
        file_submission_upload_response = requests.post(file_submission_url_json['upload_url'],
                                                        data=submission_form_data, files=files_data,
                                                        headers=Utils.canvas_api_headers())

        if file_submission_upload_response.status_code != 201:  # note: 201 Created
            print('\tERROR: unable to upload attachment file; skipping submission')
            continue

        file_submission_upload_json = file_submission_upload_response.json()
        print('\tAssociating uploaded file', file_submission_upload_json['id'], 'with new attachment comment')
        comment_association_data['comment[file_ids][]'] = [file_submission_upload_json['id']]

    comment_association_response = requests.put(user_submission_url, data=comment_association_data,
                                                headers=Utils.canvas_api_headers())
    if comment_association_response.status_code != 200:
        print('\tERROR: unable to add assignment mark/comment and associate attachment; skipping submission')
        continue

    print('\tFeedback created and associated successfully at', user_submission_url)

for key, entry in marks_map.items():
    if 'matched' not in entry:
        print('WARNING: marks file entry for', key, 'not matched to submission:', entry)
