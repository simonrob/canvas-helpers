# noinspection GrazieInspection
"""The intended use of the Canvas moderation tools is for one or more markers to initially grade submissions, and then a
moderating marker to review these, either selecting one mark as the final grade or providing their own (often naturally
an average of the existing marks). Like so many of the platform's features, this works relatively well with small
classes, but is totally impractical at larger course sizes. In addition, even with smaller classes, moderation does not
always work well when rubrics are used - any comments entered by markers whose score is not chosen as the final grade
are simply discarded. This script automates the process of averaging marks from multiple markers; and, when rubrics are
used, combines all markers' grades and feedback into a single final rubric that is released (anonymously or with marker
names attached) to the student.

The person running this script *must* be set as the assignment's moderator in order to be able to summarise and finalise
marks and feedback. In addition, it is important to be aware that finalising marks is a permanent operation, and cannot
be undone. Use of `--dry-run` is strongly recommended to preview the script's actions.

Finally, even after running this script, it is possible to revisit assessments in the submission details page (e.g.,
https://[assignment link]/submissions/[Canvas student ID]) and add a new rubric assessment that overrides all others.
Unlike those that feed into moderated marking, this will be seen as the final version, and is not edited by this tool.

Related tools:
- Create rubrics from a spreadsheet: https://community.canvaslms.com/t5/C/I/ba-p/264527 (and associated scripts at
    https://github.com/jamesjonesmath/canvancement/tree/master/rubrics)
- Analyse rubric scoring after finalising: https://community.canvaslms.com/t5/C/R/ba-p/270213
"""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2024 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2025-03-12'  # ISO 8601 (YYYY-MM-DD)

import argparse
import json
import os
import sys

import openpyxl.utils
import openpyxl.worksheet.dimensions
import requests

from canvashelpers import Args, Config, Utils


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('url', nargs=1, help='Please provide the URL of the assignment to be moderated')
    parser.add_argument('--backup-file', required=True,
                        help='Moderation of marks involves irreversible changes, so we require backing up the current '
                             'marks and feedback to a spreadsheet. Use this parameter to provide the path to an XLSX '
                             'file to use for this purpose (which will be overwritten if it exists)')
    parser.add_argument('--include-unsubmitted', action='store_true',
                        help='Students who have not made a submission for the assignment are skipped by default. Set '
                             'this option if you want to include these students (for example, when no submission is '
                             'actually expected, and the Canvas assignment is used solely to record marks). Please '
                             'note this will include any staff enrolled as students (but not the inbuilt test student)')
    parser.add_argument('--minimum-markers', type=int, default=2,  # TODO: get this from the assignment?
                        help='It can be helpful to run this script to review grading outcomes before all marking has '
                             'been completed. Use this parameter to set a minimum threshold for the number of '
                             'individual marks received for each submission (i.e., number of markers) in order to '
                             'calculate a final grade. The script will not finalise and release grades unless *all* '
                             'submissions meet this threshold')
    parser.add_argument('--identify-rubric-markers', action='store_true',
                        help='When the collated marks are released by this script, they will show as being created by '
                             'the assignment moderator. There is no way to avoid this, but if needed, set this option, '
                             'and when using a rubric the script will add individual markers\' names alongside their '
                             'scores and feedback')
    parser.add_argument('--moderator-marking', action='store_true',
                        help='The default assumption (by Canvas\'s designers) is that moderators do not mark, but '
                             'simply review marks and feedback entered by others, then select or enter a final grade. '
                             'If you (the moderator) are also a marker, set this option so that the script knows to '
                             'treat your marks as equal to any others, averaging them in the same way. Without this '
                             'option, the script will combine all markers\' feedback, but treat any mark you yourself '
                             'give as the final grade, overriding any other that are present (and the '
                             '`--minimum-markers` threshold)')
    parser.add_argument('--mark-rounding', type=float, default=0.5,
                        help='A fractional value to be used for rounding final marks. For example, 5 rounds to the '
                             'nearest 5 marks. Must be greater than 0')
    parser.add_argument('--dry-run', action='store_true',
                        help='Preview the script\'s actions without actually making any changes. Highly recommended!')
    return parser.parse_args()


def calculate_final_grade(grade_list):
    """We perform a simple average of the given list of grades, but this could be more nuanced - for example, rounding
    at the margins or other tweaks such as rejecting marks that have a high variance (by returning -1). This function
    can be adjusted to suit your needs"""
    average_grade = sum(grade_list) / len(grade_list)
    rounding_factor = 1 / args.mark_rounding  # e.g., 0.5 -> 2 to round to nearest 0.5
    average_grade = round(average_grade * rounding_factor) / rounding_factor
    return average_grade


args = Args.interactive(get_args)
ASSIGNMENT_URL = Utils.course_url_to_api(args.url[0])
ASSIGNMENT_ID = Utils.get_assignment_id(ASSIGNMENT_URL)
API_ROOT = ASSIGNMENT_URL.split('/assignments')[0]

# we need the user's details in order to differentiate between their grades (as moderator or marker) and those of others
USER_ID, user_name = Utils.get_user_details(API_ROOT, user_id='self')
if not USER_ID:
    print('ERROR: unable to get user details - did you set a valid Canvas API token in %s?' % Config.FILE_PATH)
    sys.exit()
print('Running moderation manager as %s (%d) with moderator marking: %r' % (user_name, USER_ID, args.moderator_marking))
user_map = {USER_ID: user_name}  # for use in backup file and log messages

# 1) get any associated rubric via the assignment details - if present we need rubric details before anything else
assignment_details_response = requests.get(ASSIGNMENT_URL, headers=Utils.canvas_api_headers())
if assignment_details_response.status_code != 200:
    print('ERROR: unable to get assignment details - did you set a valid Canvas API token in %s?' % Config.FILE_PATH)
    sys.exit()
assignment_details_json = assignment_details_response.json()
HAS_RUBRIC = 'rubric' in assignment_details_json
if not assignment_details_json['moderated_grading']:
    print('ERROR: assignment', ASSIGNMENT_ID, 'is not set up for moderated grading; aborting')
print('Configuring moderated assignment', ASSIGNMENT_ID, ('(with rubric)' if HAS_RUBRIC else '(no rubric)'))

# get details of the rubric - its points, criteria and link with the assignment
rubric = []
rubric_association = None
rubric_spreadsheet_map = {}
rubric_points_hidden = False
spreadsheet_headers = ['Student Number', 'Student Name', 'Marker ID', 'Marker Name', 'Overall Mark']  # backup only
if HAS_RUBRIC:
    rubric_id = assignment_details_json['rubric_settings']['id']
    print('Found rubric', rubric_id, 'associated with assignment', ASSIGNMENT_ID)

    rubric_associations_response = requests.get('%s/rubrics/%d' % (API_ROOT, rubric_id),
                                                params={'include[]': ['assignment_associations']},
                                                headers=Utils.canvas_api_headers())
    if rubric_associations_response.status_code != 200:
        print('ERROR: unable to get rubric', rubric_id, 'details; aborting')
        sys.exit()

    # extract the rubric's criteria and create columns for our backup spreadsheet
    rubric_associations_json = rubric_associations_response.json()
    for criterion in rubric_associations_json['data']:
        rubric.append({'id': criterion['id'], 'points': criterion['points'], 'description': criterion['description']})
        spreadsheet_headers.extend(['', 'Rubric: %s (max: %s; ID: %s)' % (criterion['description'], criterion['points'],
                                                                          criterion['id'])])  # '' to allow merging
        rubric_spreadsheet_map[criterion['id']] = len(spreadsheet_headers)

    # get the connection between assignment and rubric (the association)
    for association in rubric_associations_json['associations']:
        if association['association_id'] == ASSIGNMENT_ID:  # rubrics can be associated with multiple assignments
            rubric_association = association
            rubric_points_hidden = association['hide_points']
            break
    if not rubric_association:
        print('ERROR: unable to get rubric', rubric_id, 'association; aborting')
        sys.exit()

    print('Found rubric criteria:', rubric)
    print('Found rubric association', rubric_association['id'], '(points hidden: %s) -' % rubric_points_hidden,
          rubric_association)

# moderation of marks involves irreversible changes, so we back up the current state to a spreadsheet
workbook = openpyxl.Workbook()
spreadsheet = workbook.active
spreadsheet.title = 'Moderated marks (%d)' % ASSIGNMENT_ID
spreadsheet.freeze_panes = 'A2'  # set the first row as a header
spreadsheet.append(spreadsheet_headers)

dimension_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=spreadsheet)
for column in range(spreadsheet.min_column, spreadsheet.max_column + 1):  # try to size columns appropriately
    column_dimension = openpyxl.worksheet.dimensions.ColumnDimension(spreadsheet, min=column, max=column,
                                                                     width=len(spreadsheet_headers[column - 1]))
    dimension_holder[openpyxl.utils.get_column_letter(column)] = column_dimension  # stackoverflow.com/a/60801712
spreadsheet.column_dimensions = dimension_holder
header_row = spreadsheet.max_row
for column in rubric_spreadsheet_map.values():
    spreadsheet.cell(row=header_row, column=column - 1).value = spreadsheet.cell(row=header_row, column=column).value
    spreadsheet.merge_cells(start_row=header_row, end_row=header_row, start_column=column - 1, end_column=column)

# next, load the assignment's submissions as normal, but combine and average existing comments/scores
submission_list_response = Utils.get_assignment_submissions(ASSIGNMENT_URL,
                                                            includes=['provisional_grades', 'rubric_assessment'])
if not submission_list_response:
    print('ERROR: unable to retrieve submission list; aborting')
    sys.exit()

# identify and ignore the inbuilt test student
course_enrolment_response = Utils.get_course_enrolments(API_ROOT)
if not course_enrolment_response:
    print('ERROR: unable to retrieve course enrolment list; aborting')
    sys.exit()
ignored_users = [user['user_id'] for user in json.loads(course_enrolment_response)]

submission_list_json = json.loads(submission_list_response)  # note: groups mode cannot be used when enabling moderation
filtered_submission_list = Utils.filter_assignment_submissions(ASSIGNMENT_URL, submission_list_json,
                                                               include_unsubmitted=args.include_unsubmitted,
                                                               ignored_users=ignored_users, sort_entries=True)
if len(filtered_submission_list) <= 0:
    print('No valid submissions found; aborting')
    sys.exit()

final_grades = {}
skipped_submissions = set()  # we collate a list of student names whose submissions generated an error/warning
for submission in filtered_submission_list:
    submitter = Utils.get_submitter_details(ASSIGNMENT_URL, submission)
    if not submitter:
        print('\tWARNING: submitter details not found for submission; skipping:', submission)
        continue

    print('\nProcessing submission from', submitter)
    if 'provisional_grades' not in submission:
        print('\tWARNING: no provisional grades found for submission; skipping:', submitter['student_number'])
        skipped_submissions.add(submitter['student_name'])
        continue

    final_grade_id = -1
    total_score = []
    rubric_points = {}
    rubric_comments = {}
    moderator_override = False
    for rubric_criterion in rubric:
        # note: the "Remove points from rubric" option in the Canvas interface
        # doesn't actually affect this value, so we can always use points safely
        criterion_id = rubric_criterion['id']
        rubric_points[criterion_id] = []
        rubric_comments[criterion_id] = []

    # 2) for each submission, first collate the marks and rubric points/comments (if applicable) from all markers -
    # provisional grades are the grades submitted from markers but not yet selected as the final student grade
    for scorer_grade in submission['provisional_grades']:
        if scorer_grade['final']:
            final_grade_id = scorer_grade['provisional_grade_id']
            print('\tSkipping provisional grade marked as final (will be replaced by the new calculated score/rubric)',
                  '- existing details:', scorer_grade)
            # alternatively: delete the rubric assessment, but this doesn't remove the provisional grade, so ineffective
            # rubric_removal_response = requests.delete('%s/rubric_associations/%d/rubric_assessments/%d' % (
            #     API_ROOT, rubric_association['id'], marker_grade['rubric_assessments'][0]['id']),
            #     headers=Utils.canvas_api_headers())
            continue

        scorer_id = scorer_grade['scorer_id']
        overall_score = scorer_grade['score']  # note: could be empty if rubrics are used but not linked to marking

        # prepare for backing up the original entries to our spreadsheet
        if scorer_id not in user_map:
            _, scorer_name = Utils.get_user_details(API_ROOT, scorer_id)
            user_map[scorer_id] = scorer_name
        spreadsheet.append([submitter['student_number'], submitter['student_name'],
                            scorer_id, user_map[scorer_id], overall_score])

        # submissions with rubrics need a little more unpacking to get the individual points and comments
        print('\tFound provisional grade from', user_map[scorer_id], scorer_grade)
        if HAS_RUBRIC:
            if len(scorer_grade['rubric_assessments']) <= 0:
                # an overall assessment score has been entered, but no rubric details - often this is a submission given
                # 0 where there was no need to complete the rubric (e.g., absent); best to highlight for manual checking
                print('\t\tWARNING: skipping provisional grade from', user_map[scorer_id], 'with no rubric assessment')
                continue

            rubric_assessment = scorer_grade['rubric_assessments'][0]  # safe: there is only one assessment per marker
            print('\t\tFound rubric assessment from', user_map[scorer_id],
                  '(hidden points)' if rubric_points_hidden else '', '-', rubric_assessment)

            rubric_score = rubric_assessment['score']
            if rubric_score is None and not rubric_points_hidden:  # can be none if the marker/moderator has not marked
                print('\t\tWARNING: skipping rubric assessment from', user_map[scorer_id], 'with no score entered')
                continue

            row = spreadsheet.max_row
            for criterion in rubric_assessment['data']:
                criterion_id = criterion['criterion_id']

                # add rubric details to our backup spreadsheet, then build them into the final rubric (even if we have
                # overridden these points in the final mark calculation)
                position = rubric_spreadsheet_map[criterion_id]
                if 'points' in criterion:
                    spreadsheet['%s%d' % (openpyxl.utils.get_column_letter(position - 1), row)] = criterion['points']
                    rubric_points[criterion_id].append(criterion['points'])
                if criterion['comments_enabled'] and criterion['comments']:
                    spreadsheet['%s%d' % (openpyxl.utils.get_column_letter(position), row)] = criterion['comments']
                    scorer_identity = '%s: ' % user_map[scorer_id] if args.identify_rubric_markers else ''
                    rubric_comments[criterion_id].append('%s%s' % (scorer_identity, criterion['comments']))

            if moderator_override:
                continue  # we've already overridden this mark - back it up, but don't include it in the calculation

            if overall_score and overall_score != rubric_score:
                if not rubric_points_hidden:
                    print('\t\tWARNING: overall score from', user_map[scorer_id], 'of', overall_score, 'differs from',
                          'rubric score of', rubric_score, '- using overall score rather than rubric for calculation')
                total_score.append(overall_score)
            else:
                total_score.append(rubric_score)

            if scorer_id == USER_ID and not args.moderator_marking:
                # if the moderator has entered a mark and they are not themselves marking, this is always the final mark
                moderator_override = True
                total_score = [total_score[-1]]
                for rubric_criterion in rubric:  # include only the moderator's points (but everyone's comments)
                    criterion_id = rubric_criterion['id']
                    rubric_points[criterion_id] = [rubric_points[criterion_id][-1]]
                print('\t\tFound moderator override rubric - choosing final mark as', total_score[0], 'for',
                      submitter['student_number'], 'and including only comments (not points) from other rubrics')
                # break  # don't just exit - we want to back up other marks even if they are not taken into account

        else:
            if moderator_override:
                continue  # we've already overridden this mark - back it up, but don't include it in the calculation

            if scorer_id == USER_ID and not args.moderator_marking:
                # if the moderator has entered a mark and they are not themselves marking, this is always the final mark
                moderator_override = True
                total_score = [overall_score]
                print('\t\tFound moderator override grade - choosing final mark as', total_score, 'for',
                      submitter['student_number'])
                # break  # don't just exit - we want to back up other marks even if they are not taken into account
            else:
                total_score.append(overall_score)

    # 3) calculate and post the final mark (and add to spreadsheet to assist with `--dry-run` mode)
    num_scores = len(total_score)
    if num_scores <= 0:
        print('\tERROR: found submission with no valid provisional grades; skipping', submitter['student_number'])
        skipped_submissions.add(submitter['student_name'])
        continue

    total_score = sorted(total_score)
    print('\tFound a total of', num_scores, 'valid provisional grades:', total_score)
    if not moderator_override:
        if num_scores < args.minimum_markers:
            print('\tWARNING:', num_scores, 'provisional grades found for submission from', submitter['student_number'],
                  'is less than the `--minimum-markers` threshold of', args.minimum_markers, '- skipping')
            skipped_submissions.add(submitter['student_name'])
            continue

    submitter_final_grade = calculate_final_grade(total_score) if num_scores > 1 else total_score[0]
    print('\tSetting final mark from given list', total_score, 'to', submitter_final_grade)
    final_grades[submitter['canvas_user_id']] = submitter_final_grade
    if submitter_final_grade < 0:
        print('\tWARNING: unable to set final mark from', total_score, 'for', submitter['student_number'],
              '- `calculate_final_grade` returned -1; skipping')
        skipped_submissions.add(submitter['student_name'])
        continue

    print('\t%s a final mark of' % ('DRY RUN: would post' if args.dry_run else 'Posting'), submitter_final_grade, 'for',
          submitter['student_number'], '- rubric: %s, %s' % (rubric_points, rubric_comments) if HAS_RUBRIC else '')
    spreadsheet.append([submitter['student_number'], submitter['student_name'], '-1', os.path.basename(__file__),
                        submitter_final_grade])

    if HAS_RUBRIC:
        # add a new additional rubric as a summary of the individual markers' comments and scores
        new_provisional_grade_data = {'rubric_assessment[user_id]': submitter['canvas_user_id'],
                                      'rubric_assessment[assessment_type]': 'grading',
                                      # 'graded_anonymously': True,  # grading anonymously seem to do nothing in reality
                                      'provisional': True, 'final': True}  # provisional+final generates a new rubric

        row = spreadsheet.max_row
        for rubric_criterion in rubric:  # collate points/comments and add details to our backup spreadsheet
            points = rubric_points[rubric_criterion['id']]
            position = rubric_spreadsheet_map[rubric_criterion['id']]
            criteria_index = 'rubric_assessment[criterion_%s]' % rubric_criterion['id']
            if len(points) > 0:
                average_points = sum(points) / len(points)
                new_provisional_grade_data['%s[points]' % criteria_index] = average_points
                spreadsheet['%s%d' % (openpyxl.utils.get_column_letter(position - 1), row)] = average_points
            comments = rubric_comments[rubric_criterion['id']]
            new_provisional_grade_data['%s[comments]' % criteria_index] = '\n\n---\n\n'.join(comments)
            spreadsheet['%s%d' % (openpyxl.utils.get_column_letter(position), row)] = '\n\n---\n\n'.join(comments)

        if args.dry_run:
            continue

        # create (or update) the final provisional grade and rubric assessment
        rubric_link = '%s/rubric_associations/%d/rubric_assessments' % (API_ROOT, rubric_association['id'])
        if final_grade_id > -1:
            print('\tUpdating existing rubric assessment:', final_grade_id)
            rubric_method = requests.put
            rubric_link = '%s/%d' % (rubric_link, final_grade_id)
        else:
            print('\tCreating new rubric assessment')
            rubric_method = requests.post

        create_rubric_response = rubric_method(rubric_link, data=new_provisional_grade_data,
                                               headers=Utils.canvas_api_headers())
        if create_rubric_response.status_code != 200:
            print('\t\tERROR: rubric creation/update failed; skipping', create_rubric_response.text)
            skipped_submissions.add(submitter['student_name'])
            continue
        final_grade_id = create_rubric_response.json()['artifact']['provisional_grade_id']  # update if newly created

        print('\tSelecting final provisional grade rubric assessment:', final_grade_id)
        provisional_grade_selection_response = requests.put('%s/provisional_grades/%d/select' % (
            ASSIGNMENT_URL, final_grade_id), headers=Utils.canvas_api_headers())
        if provisional_grade_selection_response.status_code != 200:
            print('\t\tERROR: unable to select final provisional grade for submission; aborting. Please make sure',
                  'this tool is being run as the assignment moderator')
            skipped_submissions.add(submitter['student_name'])
            continue

workbook.save(args.backup_file)
if args.dry_run:
    print('\nDRY RUN: exiting without releasing grades')
    sys.exit()

# 4) "release" final grades - it is unclear what the purpose of this step is - it means no new marks can be entered, but
# is required in order to be able to set the final grade if this is different to the rubric calculation
print('\nReleasing moderated grades')
grades_released = False
grades_released_message = 'Assignment grades have already been published'  # Canvas's error message on failure

# we don't release grades if there are any potential issues, so need to check the release status first - if grades have
# already been released, however, we *do* proceed to finalising the grades from the calculated values (which affects
# assignments where the rubric is not used for grading, and `--mark-rounding` adjustments)
# note: the only way to detect the grade release status without attempting to actually release grades seems to be to
# submit a request for a provisional grade for any student and check the text(!) of the error message
first_student = {
    'student_id': Utils.get_submitter_details(ASSIGNMENT_URL, next(iter(filtered_submission_list)))['canvas_user_id']
}
provisional_grade_selection_response = requests.get('%s/provisional_grades/status' % ASSIGNMENT_URL,
                                                    data=first_student, headers=Utils.canvas_api_headers())
if provisional_grade_selection_response.status_code == 400 and \
        provisional_grade_selection_response.json()['message'] == grades_released_message:
    grades_released = True
print('\tModerated provisional grades have %s been released' % ('already' if grades_released else 'not'))

if skipped_submissions and not grades_released:
    print('\tERROR: not all submissions could be assigned finalised grades; aborting')
    print('\tPlease visit %s/moderate' % args.url[0], 'to check moderation, paying particular attention to students',
          'who could not be automatically moderated (see previous messages for error details):', skipped_submissions)
    print('\tClick "Release Grades" on that page if you are happy with the outputs, then re-run this tool to finalise',
          'grades. Note that grades will still not be visible to students until you actually post them')
    sys.exit()

if not grades_released:
    post_grades_response = requests.post('%s/provisional_grades/publish' % ASSIGNMENT_URL,
                                         headers=Utils.canvas_api_headers())
    if post_grades_response.status_code != 200:
        if post_grades_response.status_code == 400 and \
                post_grades_response.json()['message'] == grades_released_message:
            # somehow we got a different response here to the result of our earlier check...
            print('\tModerated provisional grades have already been released; skipping')
        else:
            print('\tERROR: unable to release provisional outcomes and select final grades')
            print('Visit %s/moderate' % args.url[0], 'and click "Release Grades", then re-run this tool to finalise',
                  'grades. Note that grades will still not be visible to students until you actually post them')
            sys.exit()
    else:
        print('\tSuccessfully released moderated grades:', post_grades_response.text)

# 5) update the final grade to reflect the average, and (in rubric mode) ensure overridden grades are included
# note that this can lead to confusion if, e.g., a high-scoring rubric is overridden with a low score or vice-versa,
# so it is normally best to just take the rubric score as the final mark (editable in the rubric's settings)
print('\nUpdating final assignment grades')
score_feedback_hint = 'See the rubric for criteria scores and a summary of marker feedback (if available)'
for submission in filtered_submission_list:
    submitter = Utils.get_submitter_details(ASSIGNMENT_URL, submission)
    if submitter['canvas_user_id'] not in final_grades:
        print('\tSkipping unmarked submission from', submitter)
        continue
    final_grade_data = {'comment[attempt]': submission['attempt'],
                        'submission[posted_grade]': final_grades[submitter['canvas_user_id']]}
    if HAS_RUBRIC:
        final_grade_data['comment[text_comment]'] = score_feedback_hint
    user_submission_url = '%s/submissions/%d' % (ASSIGNMENT_URL, submitter['canvas_user_id'])
    final_grade_response = requests.put(user_submission_url, data=final_grade_data, headers=Utils.canvas_api_headers())
    if final_grade_response.status_code != 200:
        print('\t%s' % final_grade_response.text)
        print('\tERROR: unable to finalise assignment mark/comment; skipping submission from', submitter)
        continue
