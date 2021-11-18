"""Instructure recently enforced a switch from "Classic Quizzes" to "New Quizzes". The new version has far fewer
 features (see comparison: https://docs.google.com/document/d/11nSS2EP0UpSM6dcuEFnoF-hC6lyqWbE9JSHELNmfG2A/) and
 is far harder to use, but there seems to be little interest in improving it. Critically, it is not possible to
 export bulk responses in bulk, meaning that tasks which previously took minutes now take hours for larger class
 sizes. This script uses the Canvas API to work around that limitation, exporting all responses to a spreadsheet."""
import argparse
import json
import os
import re

import configparser
import openpyxl.utils
import requests.structures

parser = argparse.ArgumentParser()
parser.add_argument('url', nargs=1,
                    help='Please pass the URL of the assignment to retrieve quiz responses for. Output will be saved '
                         'as [assignment ID].xlsx')
args = parser.parse_args()  # exits if no URL is provided

CONFIG_FILE_PATH = '%s/canvas-helpers.config' % os.path.dirname(os.path.realpath(__file__))
configparser = configparser.ConfigParser()
configparser.read(CONFIG_FILE_PATH)
CONFIG = configparser[configparser.sections()[0]]

ROOT_INSTRUCTURE_DOMAIN = 'https://%s.quiz-%s-dub-prod.instructure.com/api'
LTI_API_ROOT = ROOT_INSTRUCTURE_DOMAIN % (CONFIG['lti_institution_subdomain'], 'lti')
QUIZ_API_ROOT = ROOT_INSTRUCTURE_DOMAIN % (CONFIG['lti_institution_subdomain'], 'api')
print(LTI_API_ROOT)

ASSIGNMENT_URL = args.url[0]
ASSIGNMENT_ID = ASSIGNMENT_URL.split('/')[-1]  # used only for output spreadsheet title and filename
OUTPUT_FILE = '%s.xlsx' % ASSIGNMENT_ID
print('Exporting quiz results from assignment', ASSIGNMENT_URL, 'to', OUTPUT_FILE)
ASSIGNMENT_URL = ASSIGNMENT_URL.replace('/courses', '/api/v1/courses')

HTML_REGEX = re.compile('<.*?>')

workbook = openpyxl.Workbook()
spreadsheet = workbook.active
spreadsheet.title = 'Quiz results (%s)' % ASSIGNMENT_ID
spreadsheet.freeze_panes = 'A2'  # set the first row as a header
spreadsheet_headers = ['Student name']
spreadsheet_headers_set = False
spreadsheet_row = 2  # 1-indexed; row 1 = headers

submission_list_headers = requests.structures.CaseInsensitiveDict()
submission_list_headers['accept'] = 'application/json'
submission_list_headers['authorization'] = 'Bearer %s' % CONFIG['canvas_api_token']

# TODO: properly handle pagination: https://canvas.instructure.com/doc/api/file.pagination.html
# TODO: handle submission variants (submission_history parameter) - see: canvas.instructure.com/doc/api/submissions.html
submission_list_response = requests.get('%s/submissions/?include[]=submission_history&per_page=1000' % ASSIGNMENT_URL,
                                        headers=submission_list_headers)
if submission_list_response.status_code != 200:
    print('Error in quiz submission list retrieval - did you set a valid Canvas API token?')
    exit()

submission_list_json = json.loads(submission_list_response.text)
session_ids = []
for submission in submission_list_json:
    if 'external_tool_url' in submission:
        session_ids.append(submission['external_tool_url'].split('participant_session_id=')[1].split('&')[0])
    else:
        pass  # normally a test student
print('Loaded', len(submission_list_json), 'submission IDs:', session_ids)

token_headers = requests.structures.CaseInsensitiveDict()
token_headers['accept'] = 'application/json'
token_headers['authorization'] = ('%s' if 'Bearer ' in CONFIG['lti_bearer_token'] else 'Bearer %s') % CONFIG[
    'lti_bearer_token']  # (in case the heading 'Bearer ' is copied as well as the token itself)

for session_id in session_ids:
    print('Requesting quiz sessions for participant', session_id)
    token_response = requests.get('%s/participant_sessions/%s/grade' % (LTI_API_ROOT, session_id),
                                  headers=token_headers)
    if token_response.status_code != 200:
        print('Error in quiz session retrieval - did you set a valid browser Bearer token?')
        exit()

    attempt_json = json.loads(token_response.text)
    quiz_session_headers = requests.structures.CaseInsensitiveDict()
    quiz_session_headers['accept'] = 'application/json'
    quiz_session_headers['authorization'] = attempt_json['token']
    quiz_session_id = attempt_json['quiz_api_quiz_session_id']
    print('Loaded quiz session', quiz_session_id)

    submission_response = requests.get('%s/quiz_sessions/%d/' % (QUIZ_API_ROOT, quiz_session_id),
                                       headers=quiz_session_headers)
    if submission_response.status_code != 200:
        print('Error in quiz metadata retrieval - aborting')
        exit()

    submission_summary_json = json.loads(submission_response.text)
    student_name = submission_summary_json['metadata']['user_full_name']
    results_id = submission_summary_json['authoritative_result']['id']
    spreadsheet['A%d' % spreadsheet_row] = student_name
    print('Loaded submission summary for', student_name, '-', results_id)

    quiz_questions_response = requests.get('%s/quiz_sessions/%d/session_items' % (QUIZ_API_ROOT, quiz_session_id),
                                           headers=quiz_session_headers)
    quiz_questions_json = json.loads(quiz_questions_response.text)

    quiz_answers_response = requests.get(
        '%s/quiz_sessions/%d/results/%s/session_item_results' % (QUIZ_API_ROOT, quiz_session_id, results_id),
        headers=quiz_session_headers)
    quiz_answers_json = json.loads(quiz_answers_response.text)

    answer_number = 2  # answer 1 is always the student's name
    for question in quiz_questions_json:
        question_id = question['item']['id']
        question_type = question['item']['user_response_type']
        question_title = question['item']['title']

        if not spreadsheet_headers_set:
            spreadsheet_headers.append(question_title)
            column_letter = openpyxl.utils.get_column_letter(len(spreadsheet_headers))
            spreadsheet['%s1' % column_letter] = question_title

        print()
        print(question_title)

        current_answer = None
        for answer in quiz_answers_json:
            if answer['item_id'] == question_id:
                current_answer = answer
                break

        if current_answer:
            if question_type == 'Text':
                raw_answer = current_answer['scored_data']['value']
                if raw_answer:
                    answer_text = re.sub(HTML_REGEX, '', raw_answer)
                    print(answer_text)
                    spreadsheet[
                        '%s%d' % (openpyxl.utils.get_column_letter(answer_number), spreadsheet_row)] = answer_text

            elif question_type == 'Uuid':
                matched_answer = None
                for value in current_answer['scored_data']['value']:
                    if current_answer['scored_data']['value'][value]['user_responded']:
                        matched_answer = value

                if matched_answer:
                    for choice in question['item']['interaction_data']['choices']:
                        if choice['id'] == matched_answer:
                            answer_text = re.sub(HTML_REGEX, '', choice['item_body'])
                            print(answer_text)
                            spreadsheet['%s%d' % (
                                openpyxl.utils.get_column_letter(answer_number), spreadsheet_row)] = answer_text

            elif question_type == 'MultipleResponse':
                matched_answer = None
                for value in current_answer['scored_data']['value']:
                    for response in current_answer['scored_data']['value'][value]['value']:
                        if current_answer['scored_data']['value'][value]['value'][response]['user_responded']:
                            matched_answer = response

                if matched_answer:
                    for choice in question['item']['interaction_data']['blanks'][0]['choices']:
                        if choice['id'] == matched_answer:
                            answer_text = re.sub(HTML_REGEX, '', choice['item_body'])
                            print(answer_text)
                            spreadsheet['%s%d' % (
                                openpyxl.utils.get_column_letter(answer_number), spreadsheet_row)] = answer_text

            else:
                # TODO: handle any other response types
                print('WARNING: quiz response type', question_type, 'not currently handled - skipping')
                spreadsheet['%s%d' % (
                    openpyxl.utils.get_column_letter(answer_number),
                    spreadsheet_row)] = 'DATA MISSING - NOT YET EXPORTED'
                pass

            answer_number += 1

    spreadsheet_headers_set = True
    spreadsheet_row += 1

workbook.save(OUTPUT_FILE)
