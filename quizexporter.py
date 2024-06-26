"""Instructure recently enforced a switch from "Classic Quizzes" to "New Quizzes". The new version has far fewer
features (see comparison: https://docs.google.com/document/d/11nSS2EP0UpSM6dcuEFnoF-hC6lyqWbE9JSHELNmfG2A/) and is
far harder to use for some tasks that were previously simple, but there seems to be little interest in improving it
(see repeated forum complaints). Critically, it is not possible to export responses in bulk, meaning that tasks which
previously took minutes can now take hours for larger class sizes. This script uses the Canvas API to work around that
limitation, exporting all responses to a single spreadsheet."""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2024 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2024-04-17'  # ISO 8601 (YYYY-MM-DD)

import argparse
import json
import os
import re
import sys

import openpyxl.utils
import requests.structures

from canvashelpers import Args, Config, Utils


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('url', nargs=1,
                        help='Please provide the URL of the assignment to retrieve quiz responses for. Output will be '
                             'saved as [assignment ID].xlsx')
    parser.add_argument('--working-directory', default=None,
                        help='The location to use for output (which will be created if it does not exist). '
                             'Default: the same directory as this script')
    parser.add_argument('--overwrite', action='store_true', help='Overwrite any existing output file')
    return parser.parse_args()


args = Args.interactive(get_args)
ASSIGNMENT_URL = Utils.course_url_to_api(args.url[0])
ASSIGNMENT_ID = Utils.get_assignment_id(ASSIGNMENT_URL)  # used only for output spreadsheet title and filename

config_settings = Config.get_settings()
ROOT_INSTRUCTURE_DOMAIN = 'https://%s.quiz-%s-dub-%s.instructure.com/api'
LTI_ENVIRONMENT_TYPE = None  # auto-detected based on first submission found
LTI_INSTITUTION_SUBDOMAIN = None  # auto-detected based on first submission found
LTI_BEARER_TOKEN = config_settings['new_quiz_lti_bearer_token']
BEARER_TOKEN_ERROR_MESSAGE = ('See the configuration file instructions, and the assignment\'s SpeedGrader page: '
                              '%s/gradebook/speed_grader?assignment_id=%d') % (
                                 args.url[0].split('/assignments')[0], ASSIGNMENT_ID)
if LTI_BEARER_TOKEN.startswith('*** your'):
    print('WARNING: new_quiz_lti_bearer_token in', Config.FILE_PATH, 'seems to contain the example value.',
          BEARER_TOKEN_ERROR_MESSAGE)

OUTPUT_DIRECTORY = os.path.dirname(
    os.path.realpath(__file__)) if args.working_directory is None else args.working_directory
os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIRECTORY, '%d.xlsx' % ASSIGNMENT_ID)
if os.path.exists(OUTPUT_FILE) and not args.overwrite:
    print('ERROR: quiz result output file', OUTPUT_DIRECTORY, 'already exists - please remove or use `--overwrite`')
    sys.exit()
print('Exporting quiz results from assignment', args.url[0], 'to', OUTPUT_FILE)

HTML_REGEX = re.compile('<.*?>')  # used to filter out HTML formatting from retrieved responses

# TODO: add CSV export as an alternative (with care to handle multi-line values)
workbook = openpyxl.Workbook()
spreadsheet = workbook.active
spreadsheet.title = 'Quiz results (%d)' % ASSIGNMENT_ID
spreadsheet.freeze_panes = 'A2'  # set the first row as a header
spreadsheet_headers = ['Student number', 'Student name']
spreadsheet_headers_set = False
spreadsheet_row = 2  # 1-indexed; row 1 = headers

submission_list_response = Utils.get_assignment_submissions(ASSIGNMENT_URL)
if not submission_list_response:
    print('ERROR: unable to retrieve submission list - did you set a valid Canvas API token in %s?' % Config.FILE_PATH)
    sys.exit()

submission_list_json = json.loads(submission_list_response)
user_session_ids = []
for submission in submission_list_json:
    if 'external_tool_url' in submission:
        external_tool_url = submission['external_tool_url']
        external_tool_url_parts = external_tool_url.split('.quiz-lti-dub-')
        user_session_ids.append({'user_id': submission['user_id'],
                                 'session_id': external_tool_url.split('participant_session_id=')[1].split('&')[0]})
        if not LTI_INSTITUTION_SUBDOMAIN:
            LTI_INSTITUTION_SUBDOMAIN = external_tool_url_parts[0].split('//')[1]
        if not LTI_ENVIRONMENT_TYPE:
            LTI_ENVIRONMENT_TYPE = external_tool_url_parts[1].split('.instructure.com')[0]

    else:
        pass  # normally a test student
print('Loaded', len(user_session_ids), 'submission IDs:', user_session_ids)
LTI_API_ROOT = ROOT_INSTRUCTURE_DOMAIN % (LTI_INSTITUTION_SUBDOMAIN, 'lti', LTI_ENVIRONMENT_TYPE)
QUIZ_API_ROOT = ROOT_INSTRUCTURE_DOMAIN % (LTI_INSTITUTION_SUBDOMAIN, 'api', LTI_ENVIRONMENT_TYPE)

student_number_map = Utils.get_assignment_student_list(ASSIGNMENT_URL)
print('Loaded', len(student_number_map), 'student number mappings:', student_number_map)

token_headers = requests.structures.CaseInsensitiveDict()
token_headers['accept'] = 'application/json'
token_headers['authorization'] = ('%s' if 'Bearer ' in LTI_BEARER_TOKEN else 'Bearer %s') % \
                                 LTI_BEARER_TOKEN  # in case the heading 'Bearer ' is copied as well as the token itself

for user_session_id in user_session_ids:
    print('Requesting quiz sessions for participant', user_session_id)
    token_response = requests.get('%s/participant_sessions/%s/grade' % (LTI_API_ROOT, user_session_id['session_id']),
                                  headers=token_headers)
    if token_response.status_code != 200:
        # TODO: there doesn't seem to be an API to get this token, but is there a better alternative to the current way?
        print('ERROR: unable to load quiz session - did you set a valid new_quiz_lti_bearer_token in',
              '%s?' % Config.FILE_PATH, BEARER_TOKEN_ERROR_MESSAGE)
        sys.exit()

    # first we get a per-submission access token
    attempt_json = token_response.json()
    quiz_session_headers = requests.structures.CaseInsensitiveDict()
    quiz_session_headers['accept'] = 'application/json'
    quiz_session_headers['authorization'] = attempt_json['token']
    quiz_session_id = attempt_json['quiz_api_quiz_session_id']
    print('Loaded quiz session', quiz_session_id)

    # then a summary of the submission session and assignment overview
    submission_response = requests.get('%s/quiz_sessions/%d/' % (QUIZ_API_ROOT, quiz_session_id),
                                       headers=quiz_session_headers)
    if submission_response.status_code != 200:
        print('ERROR: unable to load quiz metadata - aborting')
        sys.exit()

    submission_summary_json = submission_response.json()
    results_id = submission_summary_json['authoritative_result']['id']
    student_name = submission_summary_json['metadata']['user_full_name']
    student_details = [s for s in student_number_map if s['user_id'] == user_session_id['user_id']]
    spreadsheet['A%d' % spreadsheet_row] = student_details[0]['student_number'] if len(student_details) == 1 else '-1'
    spreadsheet['B%d' % spreadsheet_row] = student_name
    print('Loaded submission summary for', student_name, '-', results_id)

    # then the actual quiz questions
    quiz_questions_response = requests.get('%s/quiz_sessions/%d/session_items' % (QUIZ_API_ROOT, quiz_session_id),
                                           headers=quiz_session_headers)
    quiz_questions_json = quiz_questions_response.json()

    # and finally the responses that were submitted
    quiz_answers_response = requests.get(
        '%s/quiz_sessions/%d/results/%s/session_item_results' % (QUIZ_API_ROOT, quiz_session_id, results_id),
        headers=quiz_session_headers)
    quiz_answers_json = quiz_answers_response.json()

    current_column = 3  # in our spreadsheet, column 1 is always the student's number; column 2 is always their name
    for question in quiz_questions_json:
        question_id = question['item']['id']
        question_type = question['item']['user_response_type']
        question_title = question['item']['title']

        if not spreadsheet_headers_set:
            spreadsheet_headers.append(question_title)

        print()
        print(question_title)

        current_answer = None
        for answer in quiz_answers_json:
            if answer['item_id'] == question_id:
                current_answer = answer
                break

        if current_answer:
            if question_type == 'Text':
                # text-based responses are simply recorded in the scored data
                raw_answer = current_answer['scored_data']['value']
                if raw_answer:
                    if type(raw_answer) is list:  # file upload questions
                        answer_text = raw_answer[0]['url']
                    elif type(raw_answer) is dict:  # formula questions
                        answer_text = raw_answer['user_response']
                    else:  # essay questions
                        answer_text = re.sub(HTML_REGEX, '', raw_answer)

                    print(answer_text)
                    spreadsheet[
                        '%s%d' % (openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = answer_text
                else:
                    print('ERROR: no response value found for', question_type, 'question', question_id)

            elif question_type == 'Boolean':
                for value in current_answer['scored_data']['value']:
                    if current_answer['scored_data']['value'][value]['user_responded']:
                        print(value)
                        spreadsheet[
                            '%s%d' % (openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = value
                        break

            elif question_type == 'Uuid' or question_type == 'MultipleUuid':
                # for multiple choice, multiple answer and ordering options we provide the items the user chose
                answer_parts = []
                for value in current_answer['scored_data']['value']:
                    if type(value) is dict:  # ordering question - all responses in the order given
                        selected_answer = value['user_responded']
                        answer_body = question['item']['interaction_data']['choices'][selected_answer]['item_body']
                        answer_parts.append(re.sub(HTML_REGEX, '', answer_body))
                    else:  # multiple choice or multiple answer question - include only the selected answers
                        if current_answer['scored_data']['value'][value]['user_responded']:
                            for choice in question['item']['interaction_data']['choices']:
                                if choice['id'] == value:
                                    answer_body = re.sub(HTML_REGEX, '', choice['item_body'])
                                    answer_parts.append(re.sub(HTML_REGEX, '', answer_body))
                                    break

                if len(answer_parts) > 0:
                    answer_text = ', '.join(answer_parts)
                    print(answer_text)
                    spreadsheet['%s%d' % (
                        openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = answer_text
                else:
                    print('ERROR: no response value found for', question_type, 'question', question_id)
                    spreadsheet['%s%d' % (
                        openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = ''

            elif question_type == 'MultipleResponse':
                # (note that choice lists are unhelpfully stored in a range of different formats/structures...)
                answer_parts = []
                skip_question_type = False
                for value in current_answer['scored_data']['value']:
                    if 'correct_answer' in current_answer['scored_data']['value'][value]:  # fill in the blank questions
                        answer_parts.append(
                            re.sub(HTML_REGEX, '', current_answer['scored_data']['value'][value]['user_response']))

                    else:
                        skip_question_type = True
                        break

                if len(answer_parts) > 0:
                    answer_text = ', '.join(answer_parts)
                    print(answer_text)
                    spreadsheet['%s%d' % (
                        openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = answer_text
                elif skip_question_type:
                    # TODO: this should really be separate columns for each category, but the way New Quizzes are set
                    #       up means we don't see which incorrect items were associated with which categories
                    print('WARNING: quiz response type MultipleResponse (Categorisation) not currently handled -',
                          'skipping')
                    spreadsheet['%s%d' % (
                        openpyxl.utils.get_column_letter(current_column),
                        spreadsheet_row)] = 'DATA MISSING - NOT YET EXPORTED'
                else:
                    print('ERROR: no response value found for', question_type, 'question', question_id)
                    spreadsheet['%s%d' % (
                        openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = ''

            elif question_type == 'Hash' or question_type == 'HashOfTexts':
                # TODO: we don't fully handle Hash (hot spot) or HashOfTexts (matching) questions because they are easy
                #       to mark automatically, and hard to represent in a spreadsheet except for correct/incorrect
                print('WARNING: quiz response type', question_type, 'not currently fully handled - providing only',
                      'correct or incorrect status')
                response_summary = 'Correct response: %s' % ('true' if current_answer['scored_data'][
                    'correct'] else 'false')
                print(response_summary)
                spreadsheet[
                    '%s%d' % (openpyxl.utils.get_column_letter(current_column), spreadsheet_row)] = response_summary

            else:
                # TODO: handle any other response types
                print('WARNING: quiz response type', question_type, 'not currently handled - skipping')
                spreadsheet['%s%d' % (
                    openpyxl.utils.get_column_letter(current_column),
                    spreadsheet_row)] = 'DATA MISSING - NOT YET EXPORTED'

            current_column += 1

    if not spreadsheet_headers_set:
        for header_number in range(1, len(spreadsheet_headers) + 1):  # spreadsheet indexes are 1-based
            column_letter = openpyxl.utils.get_column_letter(header_number)
            spreadsheet['%s1' % column_letter] = spreadsheet_headers[header_number - 1]
        spreadsheet_headers_set = True
    spreadsheet_row += 1

workbook.save(OUTPUT_FILE)
print('\nSaved', (spreadsheet_row - 2), 'quiz responses to', OUTPUT_FILE)
