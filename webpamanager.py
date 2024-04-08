"""Run a version of the WebPA method to calculate adjusted assignment scores based on students' ratings of self and peer
contributions to group work. Works either by generating one-question spreadsheet templates to be completed and submitted
to a separate assignment by students; or, by creating per-group quizzes to capture ratings. This script was inspired by
an offline version of the WebPA scoring system that was originally developed in R by Natalia Obukhova, Chat
Wacharamanotham and Alexander Eiselmayer.

Example usage:
1) Initialise groups on Canvas using "Create Group Set". Select "Allow self sign-up" and auto-create N groups. Ask
   students to move themselves into groups. Make sure to disable self sign-up once group membership has been finalised!
2) Create a group assignment, and select "Assign grades to each student individually" in its settings
3) Run this script in `--setup` mode to create (a) group member contribution forms or (b) quizzes, then:
    (a) Distribute these forms to group members. To streamline this process, use the `conversationcreator` script
        (though beware of filling up your personal 50MB limited `conversation attachments` folder); or, the
        `feedbackuploader` script (note that grades need to be posted before students can see comments, but posting
        only graded submissions makes *all* comments visible, which is sufficient). Create a separate individual
        assignment to gather the resulting contribution forms.
    -- or --
    (b) In quiz mode there is nothing else to do, but it may be useful to remind students about this task - see the
        `conversationcreator` script in `--groups` mode. Note that this script creates a unique quiz per group, rather
        than a single overall quiz, which allows quizzes to be customised to identify individual group members to avoid
        errors (such as misidentifying the person being rated). When setting up quizzes it can be worth waiting until
        relatively close to the desired quiz start time, so that any group membership changes are taken into account.
        If you do find that quiz group membership is outdated, see this script's `--setup-quiz-delete-existing` option
        to quickly remove these old assignments.
4) Mark the group assignment as normal
5) Use this script to calculate adjusted grades. In spreadsheet mode, use the `submissiondownloader` script to retrieve
   contribution forms. In quiz mode, submissions are retrieved automatically.
6) Use the feedbackuploader script to add the scaled marks from this script's output (webpa-final-marks.xlsx) to the
   original assignment"""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2024 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2024-04-08'  # ISO 8601 (YYYY-MM-DD)

import argparse
import contextlib
import datetime
import json
import math
import os
import random
import re
import sys
import uuid

# noinspection PyPackageRequirements
import numpy  # NumPy is a Pandas dependency, so guaranteed to be present because we require Pandas (below)
import openpyxl.styles.differential
import openpyxl.utils
# noinspection PyPackageRequirements
import pandas  # we don't list Pandas in requirements.txt to skip installing for other scripts (which do not require it)
import requests
import requests.structures

from canvashelpers import Args, Utils, Config

TIMESTAMP_FORMAT = '%Y-%m-%dT%H:%M:%SZ'  # e.g., '2024-12-31T13:30:00'

WEBPA_HEADERS = ['Respondent', 'Person', 'Student №', 'Rating', 'Comments (optional)', 'Group №']
WEBPA_QUIZ_GROUP = 'Group contribution (WebPA)'


def get_args():
    parser = argparse.ArgumentParser()
    parser_example_date = datetime.datetime.now(datetime.UTC).strftime(TIMESTAMP_FORMAT).rstrip('Z')
    parser.add_argument('group', nargs=1,
                        help='Please provide the URL of the groups page that shows the group set you wish to use for '
                             'the WebPA exercise (e.g., https://canvas.instructure.com/courses/[course-id]/groups#tab-'
                             '[set-id]). Note that Canvas does not always update the URL when switching group tabs, '
                             'so it is worth opening the desired group in a new window to be sure')
    parser.add_argument('--working-directory', default=None,
                        help='The location to use for processing and output. The script will work in a subfolder of '
                             'this directory that is named as the Canvas group set ID (see `group` parameter). When '
                             '`--setup` mode is activated and set to `spreadsheet` the given subfolder will be created '
                             'by the script (it should not already exist). When `--setup` is set to `quiz` or '
                             '`newquiz`, this parameter is only required if `--setup-quiz-export-links` is set. When '
                             '`--setup` is not specified, the use of spreadsheets is assumed, and this subfolder '
                             'should contain the individual student responses to the WebPA exercise, named as [student '
                             'number].xlsx (missing files will be treated as non-respondents). Note: see `--quiz-group'
                             '-name` for processing quiz responses. Default: the same directory as this script')
    parser.add_argument('--setup', default=None,
                        help='When this parameter is set to `quiz` or `newquiz`, the script will create Canvas quizzes '
                             '(Classic or New, respectively) to be completed by group members to rate their peers\' '
                             'contributions. If set to `spreadsheet`, the script will generate empty WebPA forms to be '
                             'distributed to group members (via, e.g., the `conversationcreator` script). If this '
                             'parameter is not set, the script will look for group members\' responses (searching in '
                             '`--working-directory`; or, alternatively, if `--quiz-group-name` is set, the named '
                             'Canvas assignment group)')
    parser.add_argument('--quiz-group-name', default=None,
                        help='When `--setup` mode is not specified, setting this parameter causes the script to look '
                             'for a Canvas assignment group with this name to load individual quizzes and responses '
                             'from. When `--setup` mode is activated and set to `quiz` or `newquiz`, this is the name '
                             'of the assignment group to place the generated quizzes within. If this parameter is not '
                             'set, the default name is `%s [current date/time]. In either case, if the assignment '
                             'group does not exist, it will be created. Note that *all* existing quizzes in the '
                             'assignment group will be assumed to be part of the current WebPA process, so if the '
                             'process is to be run multiple times (i.e., for different assignments), different group '
                             'names should be used. This parameter has no effect when `--setup` mode is activated and '
                             'set to `spreadsheet`' % WEBPA_QUIZ_GROUP)

    group_quiz = parser.add_argument_group(title='Quiz setup (see `canvashelpers.config` for additional '
                                                 'configuration). The following options only apply when `--setup` '
                                                 'mode is activated and set to `quiz` or `newquiz`')
    group_quiz.add_argument('--setup-quiz-available-from', default=None,
                            help='The date/time from which the WebPA quiz should be made available to respondents. '
                                 'This value should be specified as a timezone string - for example: %s. If not set, '
                                 'the quiz is made available immediately' % parser_example_date)
    group_quiz.add_argument('--setup-quiz-due-at', default=None,
                            help='The date/time at which WebPA quiz responses should be due. This value should be '
                                 'specified as a timezone string - for example: %s. If not set, the quiz has no due '
                                 'date' % parser_example_date)
    group_quiz.add_argument('--setup-quiz-export-links', action='store_true',
                            help='If set, the links to each contribution quiz will be exported to a spreadsheet named'
                                 '[`--quiz-group-name` value].xlsx in `--working-directory` (useful for example if '
                                 'messaging groups separately outside of this scriptz\'s operation)')

    group_spreadsheet = parser.add_argument_group(title='Spreadsheet setup. The following options only apply when '
                                                        '`--setup` mode is activated and set to `spreadsheet`')
    group_spreadsheet.add_argument('--setup-spreadsheet-template', default=None,
                                   help='An Excel template file to be used to create group members\' rating forms. '
                                        'Useful if you would like to add instructions or other content to the forms '
                                        'each group member completes. The template should already contain the '
                                        'response column headers %s as its last row. If this parameter is not set, a '
                                        'new spreadsheet will be created with these column headers.' % WEBPA_HEADERS)
    group_spreadsheet.add_argument('--setup-spreadsheet-group-output', action='store_true',
                                   help='Whether to generate a customised WebPA response form for each student '
                                        'number in the group (default); or, if set, one generic spreadsheet per group')

    group_processing = parser.add_argument_group(title='Processing. The following options only apply when `--setup` '
                                                       'mode is not active')
    group_processing.add_argument('--marks-file', required='--setup' not in ''.join(sys.argv),
                                  help='An XLSX or CSV file containing a minimum of two columns: student number (or '
                                       'group name) and original (unscaled) mark, in that order')
    group_processing.add_argument('--minimum-variance', type=float, default=0.2,
                                  help='The minimum WebPA variance level at which contribution ratings will be used to '
                                       'adjust marks. Default: 0.2')
    group_processing.add_argument('--mark-rounding', type=float, default=0.5,
                                  help='A fractional value to be used for rounding marks. For example, 5 rounds to the '
                                       'nearest 5 marks. Must be greater than 0. Default: 0.5')
    group_processing.add_argument('--maximum-mark', type=float, default=100,
                                  help='The maximum possible mark for the assignment that this exercise is being '
                                       'applied to, used to cap adjusted marks. Default: 100')
    group_processing.add_argument('--context-summaries', action='store_true',
                                  help='If set, the script will add two columns to the results spreadsheet: `Errors` '
                                       'summarises processing issues when forms were submitted but found to be '
                                       'invalid, and `Comment` provides a ready-made summary of the submission that '
                                       'can be provided to each submitting student')

    group_test = parser.add_argument_group(title='Testing, checking and review')
    group_test.add_argument('--dry-run', action='store_true',
                            help='Preview the script\'s Canvas actions without actually making any changes. Only '
                                 'applicable when creating quizzes; ignored for local file analysis/creation')
    group_test.add_argument('--setup-spreadsheet-test', action='store_true',
                            help='Tells the script to insert random responses into the generated WebPA forms '
                                 '(useful for testing). Only applicable when `--setup` mode is activated and set to '
                                 '`spreadsheet`')
    group_test.add_argument('--setup-quiz-delete-existing', action='store_true',
                            help='Delete any existing quizzes in the `--quiz-group-name` assignment group. Only '
                                 'applicable when `--setup` mode is activated and set to `quiz`. If this parameter is '
                                 'set, only this operation will be performed; quizzes will not be created, and other'
                                 'configuration options except for `--dry-run` will be ignored')
    return parser.parse_args()


class GroupResponseProcessor:
    @staticmethod
    def setup_spreadsheets(groups):
        if TEMPLATE_FILE:
            response_template_workbook = openpyxl.load_workbook(TEMPLATE_FILE)
            response_template_sheet = response_template_workbook[response_template_workbook.sheetnames[0]]
        else:
            response_template_workbook = openpyxl.Workbook()
            response_template_sheet = response_template_workbook.active
            response_template_sheet.title = 'WebPA response form'
            response_template_sheet.append(WEBPA_HEADERS)
        initial_max_rows = response_template_sheet.max_row

        # noinspection SpellCheckingInspection
        thin_border = openpyxl.styles.borders.Border(
            left=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
            right=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
            top=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
            bottom=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA')
        )
        output_count = 0
        for group_key in sorted(groups):
            for group_member in groups[group_key]:
                response_template_sheet.append(
                    [None, group_member['student_name'], group_member['student_number'], None, None, group_key])

            if TEMPLATE_FILE:  # highlight the part of the template that needs to be completed
                for response_row in response_template_sheet.iter_rows(min_row=initial_max_rows):
                    for cell in [response_row[0], response_row[1], response_row[2], response_row[5]]:
                        cell.fill = openpyxl.styles.PatternFill(start_color='00E7E6E6', end_color='00E7E6E6',
                                                                fill_type='solid')
                    for cell in response_row:
                        cell.border = thin_border

            if args.setup_spreadsheet_group_output:
                # just a generic form for the whole group to complete (and select their own row manually)
                response_template_workbook.save(os.path.join(WORKING_DIRECTORY, 'group-%d.xlsx' % group_key))
                output_count += 1

            else:
                # create a personalised form for each group member (with their own row pre-selected)
                for group_member in groups[group_key]:
                    for response_row in response_template_sheet.iter_rows(min_row=initial_max_rows + 1, max_col=4):
                        response_row[0].value = None
                        if response_row[2].value == group_member['student_number']:
                            response_row[0].value = '✔'
                            response_row[0].alignment = openpyxl.styles.Alignment(horizontal='center')

                        if args.setup_spreadsheet_test:
                            response_row[3].value = random.randint(1, 5)
                            print('WARNING: TEST MODE is active; generating sample response data:',
                                  response_row[3].value)

                    response_template_workbook.save(
                        os.path.join(WORKING_DIRECTORY, '%s.xlsx' % group_member['student_number']))
                    output_count += 1

            # reset for next group
            response_template_sheet.delete_rows(initial_max_rows + 1,
                                                response_template_sheet.max_row - initial_max_rows)
        print('Successfully generated', output_count, 'WebPA forms to', WORKING_DIRECTORY)

    @staticmethod
    def setup_quizzes(groups, assignment_group_id):
        # the quiz can be customised in the canvashelpers.config file
        config_settings = Config.get_settings()

        # for ease, we build the quiz link list every time, and just don't save it if not required
        quiz_link_workbook = openpyxl.Workbook()
        quiz_link_workbook_sheet = quiz_link_workbook.active
        quiz_link_workbook_sheet.title = 'WebPA quiz links'
        quiz_link_workbook_sheet.freeze_panes = 'A2'  # set the first row as a header
        quiz_link_workbook_sheet.append(['Group name', 'Quiz link'])

        output_count = 0
        for group_key in sorted(groups):
            # each group has a separate quiz that is only accessible to that group's members
            print('\nCreating WebPA quiz for student group', group_key, '(%s members)' % len(groups[group_key]))
            quiz_configuration = {
                'quiz[title]': '%s [%s]' % (config_settings['webpa_quiz_title'], groups[group_key][0]['group_name']),
                'quiz[description]': config_settings['webpa_quiz_description'],
                'quiz[quiz_type]': 'graded_survey',
                'quiz[assignment_group_id]': assignment_group_id,
                'quiz[show_correct_answers]': 'false',  # note: must be a string not a boolean
                'quiz[only_visible_to_overrides]': True
            }
            if args.setup_quiz_available_from:
                quiz_configuration['quiz[unlock_at]'] = args.setup_quiz_available_from
            if args.setup_quiz_due_at:
                quiz_configuration['quiz[due_at]'] = args.setup_quiz_due_at
                quiz_configuration['quiz[lock_at]'] = args.setup_quiz_due_at

            if args.dry_run:
                print('\tDRY RUN: skipping creation of new quiz:', quiz_configuration['quiz[title]'])
                current_quiz_id = -1
                current_quiz_assignment_id = -1
            else:
                quiz_creation_response = requests.post('%s/quizzes' % COURSE_URL,
                                                       data=quiz_configuration, headers=Utils.canvas_api_headers())
                if quiz_creation_response.status_code != 200:
                    print('\tERROR: unable to create quiz for group', group_key, ':', quiz_creation_response.text,
                          '- aborting')
                    sys.exit()

                quiz_creation_response_json = quiz_creation_response.json()
                current_quiz_id = quiz_creation_response_json['id']
                current_quiz_assignment_id = quiz_creation_response_json['assignment_id']
                print('\tCreated quiz', quiz_configuration['quiz[title]'], '- ID', current_quiz_id,
                      'with assignment ID', current_quiz_assignment_id)

            # each member has a separate contribution question
            question_position = 0
            for member_id, group_member in enumerate(groups[group_key]):
                student_description = '%s (%s)' % (group_member['student_name'], group_member['student_number'])
                question_position = member_id + 1  # uses 1-based indexing
                quiz_question_configuration = {
                    'question[question_name]': group_member['student_number'],
                    'question[question_text]': config_settings['webpa_rating_question_description'].replace(
                        '{group member}', student_description),
                    'question[question_type]': 'numerical_question',
                    'question[position]': question_position,
                    'question[answers][0][answer_range_start]': '1',  # note incorrect API documentation (ditto below)
                    'question[answers][0][answer_range_end]': '5',
                    'question[answers][0][numerical_answer_type]': 'range_answer'
                }

                if args.dry_run:
                    print('\tDRY RUN: skipping creation of new quiz question:',
                          quiz_question_configuration['question[question_name]'])
                else:
                    quiz_question_response = requests.post('%s/quizzes/%s/questions' % (COURSE_URL, current_quiz_id),
                                                           data=quiz_question_configuration,
                                                           headers=Utils.canvas_api_headers())
                    if quiz_question_response.status_code != 200:
                        print('\tERROR: unable to create question',
                              quiz_question_configuration['question[question_name]'],
                              'for quiz:', quiz_question_response.text)
                        sys.exit()
                    print('\tCreated quiz question:', quiz_question_configuration['question[question_name]'])

            # at the end we ask for any general comments - in most cases these are not used, but students often like to
            # be able to provide this (note: if left empty Canvas warns the student, hence the prompt to enter "None")
            quiz_question_configuration = {
                'question[question_name]': 'Comments (optional)',  # not currently customisable as Canvas hides this
                'question[question_text]': config_settings['webpa_comment_question_description'],
                'question[question_type]': 'short_answer_question',
                'question[position]': question_position + 1
            }

            if args.dry_run:
                print('\tDRY RUN: skipping creation of general comments quiz question:',
                      quiz_question_configuration['question[question_name]'])
            else:
                quiz_question_response = requests.post('%s/quizzes/%s/questions' % (COURSE_URL, current_quiz_id),
                                                       data=quiz_question_configuration,
                                                       headers=Utils.canvas_api_headers())
                if quiz_question_response.status_code != 200:
                    print('\tERROR: unable to create general comments question',
                          quiz_question_configuration['question[question_name]'], 'for quiz:',
                          quiz_question_response.text)
                    sys.exit()
                print('\tCreated general comments quiz question:',
                      quiz_question_configuration['question[question_name]'])

            # an update push is required after editing questions
            quiz_configuration['quiz[notify_of_update]'] = 'false'
            quiz_configuration['quiz[published]'] = True
            if args.dry_run:
                print('\tDRY RUN: skipping update push for quiz', quiz_configuration['quiz[title]'])
            else:
                quiz_update_response = requests.put('%s/quizzes/%s' % (COURSE_URL, current_quiz_id),
                                                    data=quiz_configuration, headers=Utils.canvas_api_headers())
                if quiz_update_response.status_code != 200:
                    print('\tERROR: unable to update quiz', quiz_configuration['quiz[title]'], ':',
                          quiz_update_response.text, '- aborting')
                    sys.exit()
                print('\tPushed update for quiz', quiz_configuration['quiz[title]'])

            # finally, configure access so that only this group's members can see and respond to this particular quiz
            current_group_canvas_ids = [student['student_canvas_id'] for student in groups[group_key]]
            GroupResponseProcessor.configure_quiz_access(current_quiz_assignment_id, current_group_canvas_ids)

            if not args.dry_run:
                quiz_link = '%s/quizzes/%s' % (COURSE_URL.replace('/api/v1', ''), current_quiz_id)
                quiz_link_workbook_sheet.append([groups[group_key][0]['group_name'], quiz_link])
                print('\tFinished configuring quiz at', quiz_link)
                if args.setup_quiz_export_links:
                    pass
            output_count += 1

        if args.setup_quiz_export_links:
            quiz_link_file = os.path.join(WORKING_DIRECTORY, '%s.xlsx' % args.quiz_group_name)
            print('%s quiz links to' % ('DRY RUN: skipping saving' if args.dry_run else 'Saving'), quiz_link_file)
            if not args.dry_run:
                quiz_link_workbook.save(quiz_link_file)

        print('Finished processing', output_count, 'groups')
        return

    @staticmethod
    def setup_new_quizzes(groups, assignment_group_id):
        # the quiz can be customised in the canvashelpers.config file
        config_settings = Config.get_settings()

        # for ease, we build the quiz link list every time, and just don't save it if not required
        quiz_link_workbook = openpyxl.Workbook()
        quiz_link_workbook_sheet = quiz_link_workbook.active
        quiz_link_workbook_sheet.title = 'WebPA new quiz links'
        quiz_link_workbook_sheet.freeze_panes = 'A2'  # set the first row as a header
        quiz_link_workbook_sheet.append(['Group name', 'New quiz link'])

        output_count = 0
        for group_key in sorted(groups):
            # each group has a separate quiz that is only accessible to that group's members
            print('\nCreating WebPA new quiz for student group', group_key, '(%s members)' % len(groups[group_key]))
            quiz_configuration = {
                'quiz[title]': '%s [%s]' % (config_settings['webpa_quiz_title'], groups[group_key][0]['group_name']),
                'quiz[instructions]': config_settings['webpa_quiz_description'],
                'quiz[assignment_group_id]': assignment_group_id,
                'quiz[quiz_settings][result_view_settings][result_view_restricted]': True,  # note documentation typo
                'quiz[quiz_settings][result_view_settings][display_points_awarded]': 'false',  # note: must be a string
                'quiz[quiz_settings][result_view_settings][display_points_possible]': 'false',
                'quiz[quiz_settings][result_view_settings][display_items]': 'false'
            }
            if args.setup_quiz_available_from:
                quiz_configuration['quiz[unlock_at]'] = args.setup_quiz_available_from
            if args.setup_quiz_due_at:
                quiz_configuration['quiz[due_at]'] = args.setup_quiz_due_at
                quiz_configuration['quiz[lock_at]'] = args.setup_quiz_due_at

            if args.dry_run:
                print('\tDRY RUN: skipping creation of new quiz:', quiz_configuration['quiz[title]'])
                current_quiz_id = -1
            else:
                quiz_creation_response = requests.post('%s/quizzes' % GroupResponseProcessor.new_quiz_api(COURSE_URL),
                                                       data=quiz_configuration, headers=Utils.canvas_api_headers())
                if quiz_creation_response.status_code != 200:
                    print('\tERROR: unable to create new quiz for group', group_key, ':', quiz_creation_response.text,
                          '- aborting')
                    sys.exit()

                quiz_creation_response_json = quiz_creation_response.json()
                current_quiz_id = quiz_creation_response_json['id']
                print('\tCreated new quiz', quiz_configuration['quiz[title]'], '- ID', current_quiz_id)

            # each member has a separate contribution question
            question_position = 0
            for member_id, group_member in enumerate(groups[group_key]):
                student_description = '%s (%s)' % (group_member['student_name'], group_member['student_number'])
                question_position = member_id + 1  # uses 1-based indexing
                quiz_question_configuration = {
                    'item': {
                        'entry_type': 'Item',
                        'position': question_position,
                        'entry': {
                            'interaction_type_slug': 'choice',
                            'title': group_member['student_number'],
                            'item_body': config_settings['webpa_rating_question_body'].replace('{group member}',
                                                                                               student_description),
                        }
                    }
                }

                answer_uuid = None
                interaction_data = []
                for i in range(5):
                    answer_rating = i + 1
                    answer_uuid = str(uuid.uuid4())
                    interaction_data.append({
                        'id': answer_uuid,
                        'position': answer_rating,
                        'itemBody': '<p><b>%d</b>: %s</p>' % (  # to guarantee starting with the correct rating number
                            answer_rating, config_settings['webpa_rating_question_choice_%d' % answer_rating])
                    })
                quiz_question_configuration['item']['entry']['interaction_data'] = {'choices': interaction_data}
                quiz_question_configuration['item']['entry']['scoring_algorithm'] = 'Equivalence'
                quiz_question_configuration['item']['entry']['scoring_data'] = {'value': answer_uuid}

                if args.dry_run:
                    print('\tDRY RUN: skipping creation of new quiz question:',
                          quiz_question_configuration['question[question_name]'])
                else:
                    quiz_question_response = requests.post(
                        '%s/quizzes/%s/items' % (GroupResponseProcessor.new_quiz_api(COURSE_URL), current_quiz_id),
                        json=quiz_question_configuration, headers=Utils.canvas_api_headers())
                    if quiz_question_response.status_code != 200:
                        print('\tERROR: unable to create question',
                              quiz_question_configuration['item']['entry']['title'], 'for quiz:',
                              quiz_question_response.text)
                        sys.exit()
                    print('\tCreated new quiz question:', quiz_question_configuration['item']['entry']['title'])

            # at the end we ask for any general comments - in most cases these are not used, but students often like to
            # be able to provide this (note: if left empty Canvas warns the student, hence the prompt to enter "None")
            quiz_question_configuration = {
                'item[entry_type]': 'Item',
                'item[position]': question_position + 1,
                'item[entry][interaction_type_slug]': 'essay',
                'item[entry][title]': 'Comments (optional)',  # not currently customisable as Canvas hides this
                'item[entry][item_body]': config_settings['webpa_comment_question_description'],
                'item[entry][interaction_data][rce]': 'false',  # note: must be a string not a boolean
                'item[entry][scoring_algorithm]': 'None',
                'item[entry][scoring_data][value]': ''
            }

            if args.dry_run:
                print('\tDRY RUN: skipping creation of general comments new quiz question:',
                      quiz_question_configuration['question[question_name]'])
            else:
                quiz_question_response = requests.post(
                    '%s/quizzes/%s/items' % (GroupResponseProcessor.new_quiz_api(COURSE_URL), current_quiz_id),
                    data=quiz_question_configuration,
                    headers=Utils.canvas_api_headers())
                if quiz_question_response.status_code != 200:
                    print('\tERROR: unable to create general comments question',
                          quiz_question_configuration['item[entry][title]'], 'for quiz:',
                          quiz_question_response.text)
                    sys.exit()
                print('\tCreated general comments new quiz question:',
                      quiz_question_configuration['item[entry][title]'])

            # publish via the assignments (rather than New Quizzes) API
            assignment_configuration = {
                'assignment[published]': True,
                'assignment[only_visible_to_overrides]': True,
                'assignment[omit_from_final_grade]': True,
                'assignment[hide_in_gradebook]': True
            }
            if args.dry_run:
                print('\tDRY RUN: skipping update push for new quiz', quiz_configuration['quiz[title]'])
            else:
                quiz_update_response = requests.put('%s/assignments/%s' % (COURSE_URL, current_quiz_id),
                                                    data=assignment_configuration, headers=Utils.canvas_api_headers())
                if quiz_update_response.status_code != 200:
                    print('\tERROR: unable to update new quiz', quiz_configuration['quiz[title]'], ':',
                          quiz_update_response.text, '- aborting')
                    sys.exit()
                print('\tPushed update for new quiz', quiz_configuration['quiz[title]'])

            # finally, configure access so that only this group's members can see and respond to this particular quiz
            current_group_canvas_ids = [student['student_canvas_id'] for student in groups[group_key]]
            GroupResponseProcessor.configure_quiz_access(current_quiz_id, current_group_canvas_ids)

            if not args.dry_run:
                quiz_link = '%s/assignments/%s' % (COURSE_URL.replace('/api/v1', ''), current_quiz_id)
                quiz_link_workbook_sheet.append([groups[group_key][0]['group_name'], quiz_link])
                print('\tFinished configuring new quiz at', quiz_link)
                if args.setup_quiz_export_links:
                    pass
            output_count += 1

        if args.setup_quiz_export_links:
            quiz_link_file = os.path.join(WORKING_DIRECTORY, '%s.xlsx' % args.quiz_group_name)
            print('%s new quiz links to' % ('DRY RUN: skipping saving' if args.dry_run else 'Saving'), quiz_link_file)
            if not args.dry_run:
                quiz_link_workbook.save(quiz_link_file)

        print('Finished processing', output_count, 'groups')
        return

    @staticmethod
    def new_quiz_api(original_api_url):
        return original_api_url.replace('/api/v1/', '/api/quiz/v1/')  # hosted in a different location, bizarrely

    @staticmethod
    def configure_quiz_access(current_quiz_id, current_group_canvas_ids):
        access_override_configuration = {'assignment_override[student_ids][]': current_group_canvas_ids}
        if args.setup_quiz_available_from:
            access_override_configuration['assignment_override[unlock_at]'] = args.setup_quiz_available_from
        if args.setup_quiz_due_at:
            access_override_configuration['assignment_override[due_at]'] = args.setup_quiz_due_at
            access_override_configuration['assignment_override[lock_at]'] = args.setup_quiz_due_at
        if args.dry_run:
            print('\tDRY RUN: skipping quiz assignment access configuration for Canvas users:',
                  current_group_canvas_ids, 'available from', args.setup_quiz_available_from, 'and due at',
                  args.setup_quiz_due_at)
        else:
            access_override_response = requests.post(
                '%s/assignments/%s/overrides' % (COURSE_URL, current_quiz_id),
                data=access_override_configuration, headers=Utils.canvas_api_headers())
            if access_override_response.status_code != 201:  # note 201 Created not 200 OK
                print('\tERROR: unable to configure quiz assignment access for Canvas users', current_group_canvas_ids,
                      ':', access_override_response.text, '- aborting')
            print('\tConfigured quiz assignment access for Canvas users', current_group_canvas_ids, 'available from',
                  args.setup_quiz_available_from, 'and due at', args.setup_quiz_due_at)

    @staticmethod
    def get_assignment_group_id(group_name):
        assignment_group_response = requests.get('%s/assignment_groups' % COURSE_URL,
                                                 headers=Utils.canvas_api_headers())

        if assignment_group_response.status_code != 200:
            return None

        assignment_group_response_json = assignment_group_response.json()
        for group_properties in assignment_group_response_json:
            if group_properties['name'] == group_name:
                return group_properties['id']

    @staticmethod
    def create_assignment_group(new_group_name):
        group_creation_response = requests.post('%s/assignment_groups' % COURSE_URL,
                                                data={'name': new_group_name},
                                                headers=Utils.canvas_api_headers())
        if group_creation_response.status_code != 200:
            print('\tERROR: unable to create assignment group; aborting')
            sys.exit()

        return group_creation_response.json()['id']

    @staticmethod
    def get_spreadsheets(groups, summary_sheet):
        response_files = [f for f in os.listdir(WORKING_DIRECTORY) if re.match(r'\d+\.xlsx?', f)]
        expected_submissions = [member['student_number'] for group_key in groups.values() for member in group_key]
        errors = {}
        skipped_files = []
        for file in response_files:
            invalid_file = False
            response_workbook = openpyxl.load_workbook(os.path.join(WORKING_DIRECTORY, file))
            response_sheet = response_workbook[response_workbook.sheetnames[0]]
            # response_sheet.column_dimensions['C'].number_format = '@'  # force column format to text - doesn't work

            found_header_row = False
            valid_members = []
            expected_rater = file.split('.')[0]
            if expected_rater not in expected_submissions:
                print('WARNING: skipping unexpected form', file)
                skipped_files.append(file)
                continue

            current_group = None
            current_rater = None
            current_responses = []
            current_errors = []
            current_total = 0
            found_members = []
            for response_row in response_sheet.iter_rows(max_col=6):
                cells = [c.value for c in response_row]
                if cells == WEBPA_HEADERS:
                    found_header_row = True
                    continue
                if all(v is None for v in cells):
                    continue  # sometimes openpyxl produces hundreds of empty rows at the end of a table - ignore

                if found_header_row:
                    if not cells[2]:
                        continue  # sometimes xlsx files contain empty rows after content - ignore
                    cells[2] = str(cells[2]).split('.')[0]  # make sure student number is treated as a string
                    found_members.append(cells[2])  # so we can check that all expected members are present

                    if not current_group:
                        current_group = cells[5]
                        valid_members = [g['student_number'] for g in groups[current_group]]

                    # validate the submitted data against Canvas group membership
                    ignored_rating = False
                    if cells[0]:  # note that we accept any content, not just the '✔' we ask for
                        if not current_rater and cells[2] == expected_rater:
                            current_rater = cells[2]
                        else:
                            current_errors.extend(
                                [e for e in ['Incorrect or multiple respondents selected'] if e not in current_errors])
                            invalid_file = True
                    if cells[2] not in valid_members:
                        ignored_rating = True  # not necessarily invalid - see membership checks below
                    if cells[5] != current_group:
                        current_errors.append('Invalid group number (%s)' % cells[5])
                        invalid_file = True
                    if cells[3] is None or type(cells[3]) not in [int, float]:
                        current_errors.append(
                            '%s rating %s' % ('Own' if cells[2] == current_rater else 'Member %s' % cells[2],
                                              'invalid (\'%s\')' % cells[3] if cells[3] else 'missing'))
                        invalid_file = True

                    if not (invalid_file or ignored_rating):
                        bounded_score = round(max(min(cells[3], 5), 1))  # don't allow scores outside 1-5 (int) range
                        if bounded_score != cells[3]:
                            current_errors.append('Rating %s for %s is outside of range 1-5 (rounded to %d)' % (
                                cells[3], cells[2], bounded_score))

                        current_responses.append([None, cells[2], bounded_score, None, current_group])
                        current_total += bounded_score

            if current_group:
                sorted_found = sorted(found_members)
                sorted_expected = sorted(valid_members)
                if sorted_found != sorted_expected:
                    members_missing = set(sorted_expected) - set(sorted_found)
                    if members_missing:
                        current_errors.append('Group member(s) missing: %s' % ', '.join(members_missing))
                        invalid_file = True
                    members_added = set(sorted_found) - set(sorted_expected)
                    if members_added:  # note: this can have legitimate explanations - e.g., group members withdrawing
                        current_errors.append('Non-group member(s) found: %s – ignoring' % ', '.join(members_added))

            if not current_rater:
                if not found_header_row:
                    current_errors.append('Incorrect (or edited example) rating form has been used')
                else:
                    current_errors.append('Own name indicator missing')
                invalid_file = True
            if current_errors:
                errors[expected_rater] = current_errors

            if not invalid_file:
                if current_errors:
                    print('WARNING: form data required corrections', file, '-', current_errors)
                for response in current_responses:
                    response[0] = current_rater
                    response[3] = response[2] / current_total
                    summary_sheet.append(response)
            else:
                print('ERROR: skipping invalid form', file, '-', current_errors)
                skipped_files.append(file)

        respondents = [f.split('.')[0] for f in response_files if f not in skipped_files]  # without invalid files
        invalid = [f.split('.')[0] for f in skipped_files]
        return respondents, invalid, errors

    @staticmethod
    def get_quizzes(groups, summary_sheet, quiz_group_name):
        expected_submissions = [member['student_number'] for group_key in groups.values() for member in group_key]
        respondents = []
        invalid = []
        errors = {}

        # first get all quizzes within the given assignment group
        assignment_group_id = GroupResponseProcessor.get_assignment_group_id(quiz_group_name)
        if not assignment_group_id:
            print('ERROR: unable to find quiz group name', quiz_group_name, '- aborting')
            sys.exit()

        assignment_list_response = Utils.canvas_multi_page_request(
            '%s/assignment_groups/%s/assignments' % (COURSE_URL, assignment_group_id), type_hint='assignment list')
        if not assignment_list_response:
            print('\tERROR: unable to get assignment list response; aborting')
            sys.exit()

        assignment_list_response_json = json.loads(assignment_list_response)
        for quiz in assignment_list_response_json:
            if 'quiz_id' not in quiz:
                # avoid having to specify quiz type for analysis by detecting the type of the first submission
                print('WARNING: found new quiz assignment', quiz['id'], '- switching to new quizzes mode')
                return GroupResponseProcessor.get_new_quizzes(groups, expected_submissions, summary_sheet,
                                                              assignment_list_response_json)

            quiz_id = quiz['quiz_id']
            print('\nFound quiz ID', quiz_id, '-', quiz['name'], 'with assignment ID', quiz['id'], 'due at',
                  quiz['due_at'])

            current_group = int(quiz['name'].split('[')[-1].rstrip(']').split(' ')[-1])
            valid_members = [g['student_number'] for g in groups[current_group]]
            print('\tIdentified group', current_group, 'with expected members', valid_members)

            # then all quiz questions
            question_student_map = {}
            quiz_question_response = requests.get('%s/quizzes/%s/questions' % (COURSE_URL, quiz_id),
                                                  headers=Utils.canvas_api_headers())
            if quiz_question_response.status_code != 200:
                print('\tERROR: unable to get quiz questions for quiz', quiz_id, '- aborting:',
                      quiz_question_response.text)
                sys.exit()

            quiz_question_response_json = quiz_question_response.json()
            print('\tFound', end=' ')
            for question in quiz_question_response_json:
                question_id = question['id']
                question_name = question['question_name']
                rating_question = False
                if question_name.isdigit():
                    rating_question = True
                    question_student_map[question_id] = question_name
                print('%s question' % ('rating' if rating_question else 'comments'), question_id,
                      'titled:', question_name, end='; ')
            print()

            # then all submissions for that quiz
            quiz_submission_response = requests.get('%s/quizzes/%s/submissions' % (COURSE_URL, quiz_id),
                                                    headers=Utils.canvas_api_headers())
            if quiz_submission_response.status_code != 200:
                print('\tERROR: unable to get quiz submissions for quiz', quiz_id, '- aborting:',
                      quiz_submission_response.text)
                sys.exit()

            quiz_submission_response_json = quiz_submission_response.json()
            current_quiz_submission = quiz_submission_response_json['quiz_submissions']
            if len(current_quiz_submission) <= 0:
                print('\tNo submissions found for quiz', quiz_id, '- skipping')
                continue

            for submission in current_quiz_submission:
                print('\tLoading quiz', quiz_id, 'submission:', submission['id'])

                # then a single submission's details
                quiz_submission_individual_response = requests.get(
                    '%s/quizzes/%s/submissions/%s' % (COURSE_URL, quiz_id, submission['id']),
                    params={'include[]': ['submission', 'quiz', 'user', 'submission_history']},
                    headers=Utils.canvas_api_headers())
                if quiz_submission_individual_response.status_code != 200:
                    print('\t\tERROR: unable to get individual quiz response', submission['id'], '- aborting:',
                          quiz_submission_individual_response.text)
                    sys.exit()

                # an array is returned, but we expect (and asked for) only one result, so this is okay
                submission_summary = quiz_submission_individual_response.json()['submissions'][0]
                submission_from = submission_summary['user']
                current_rater = submission_from['login_id']
                current_rater_name = submission_from['name']
                if current_rater not in expected_submissions:
                    print('\t\tWARNING: skipping unexpected form from student not in any group:', current_rater)
                    invalid.append(current_rater)
                    continue
                if current_rater not in valid_members:
                    print('\t\tWARNING: skipping unexpected form from student not in current group:', current_rater)
                    invalid.append(current_rater)
                    continue
                if submission_summary['workflow_state'] not in ['complete', 'graded', 'pending_review']:
                    print('\t\tWARNING: skipping empty or partly-complete form from', current_rater, '-',
                          submission_summary)
                    invalid.append(current_rater)
                    continue

                # date is oddly sometimes missing even if previously set
                due_date = quiz['due_at'] or submission_summary['cached_due_date']
                if due_date:
                    if (datetime.datetime.strptime(submission_summary['submitted_at'], TIMESTAMP_FORMAT) >
                            datetime.datetime.strptime(due_date, TIMESTAMP_FORMAT)):
                        print('\t\tWARNING: skipping late rating submission from', current_rater, '- submitted at',
                              submission_summary['submitted_at'], 'but due at', due_date)
                        invalid.append(current_rater)
                        continue
                print('\t\tFound submission from', current_rater_name, '- Canvas ID:', submission_from['id'],
                      '; student number:', current_rater)

                current_responses = []
                current_errors = []
                current_total = 0
                found_members = []
                invalid_response = False

                # we only allow one submission to the rating quiz, but just in case, take the first valid one found
                submission_answers = None
                for history_entry in submission_summary['submission_history']:
                    if history_entry['workflow_state'] in ['complete', 'graded'] and 'submission_data' in history_entry:
                        submission_answers = history_entry
                if not submission_answers:
                    print('\t\tWARNING: skipping unanswered or partly-complete form from ', current_rater, '-',
                          submission_summary)
                    invalid.append(current_rater)
                    continue
                submission_answers = submission_summary['submission_history'][0]['submission_data']
                for answer in submission_answers:
                    answer_value = answer['text']
                    if answer['question_id'] in question_student_map:
                        rated_student = question_student_map[answer['question_id']]
                        found_members.append(rated_student)
                        print('\t\tRating from', current_rater_name, 'for', rated_student, ':', answer_value)

                        # validate the submitted data against Canvas group membership (remembering this may change)
                        if rated_student not in valid_members:
                            print('\t\tWARNING: Ignoring rating by', current_rater, 'of non member', rated_student)
                            continue

                        try:
                            original_score = float(answer_value)
                            bounded_score = round(max(min(original_score, 5), 1))  # only permit scores 1-5 (int)
                            if bounded_score != original_score:
                                current_errors.append('Rating %s for %s is outside of range 1-5 (rounded to %d)' % (
                                    original_score, rated_student, bounded_score))

                            current_responses.append([current_rater, rated_student, bounded_score, None, current_group])
                            current_total += bounded_score

                        except ValueError:
                            current_errors.append('%s rating %s' % (
                                'Own' if rated_student == current_rater else 'Member %s' % rated_student,
                                'invalid (\'%s\')' % answer_value if answer_value else 'missing'))
                            invalid_response = True

                    elif answer_value and answer_value.lower().strip() != 'none':
                        print('\t\tWARNING: Comments from', current_rater_name, ':', answer_value)

                # finally, check for errors and collate responses
                # noinspection DuplicatedCode
                if current_group:
                    sorted_found = sorted(found_members)
                    sorted_expected = sorted(valid_members)
                    if sorted_found != sorted_expected:
                        members_missing = set(sorted_expected) - set(sorted_found)
                        if members_missing:
                            current_errors.append('Group member(s) missing: %s' % ', '.join(members_missing))
                            invalid_response = True
                        members_added = set(sorted_found) - set(sorted_expected)
                        if members_added:  # note: this can have legitimate explanations - e.g., members withdrawing
                            current_errors.append(
                                'Non-group member(s) found: %s – ignoring' % ', '.join(str(m) for m in members_added))

                if current_errors:
                    errors[current_rater] = current_errors

                if not invalid_response:
                    respondents.append(current_rater)
                    if current_errors:
                        print('\tWARNING: form data required corrections', current_rater, '-', current_errors)
                    for response in current_responses:
                        response[3] = response[2] / current_total
                        summary_sheet.append(response)
                else:
                    print('\tERROR: skipping invalid form from', current_rater, '-', current_errors)
                    invalid.append(current_rater)

        return respondents, invalid, errors

    @staticmethod
    def get_new_quizzes(groups, expected_submissions, summary_sheet, assignment_list_response_json):
        # frustratingly, much of this code needs to be duplicated from the quiz exporter script due to the lack of a new
        # quizzes response API (note also we need to use Canvas IDs far more because New Quizzes hide student numbers)
        respondents = []
        invalid = []
        errors = {}

        config_settings = Config.get_settings()
        root_instructure_domain = 'https://%s.quiz-%s-dub-%s.instructure.com/api'
        lti_environment_type = None  # auto-detected based on first submission found
        lti_institution_subdomain = None  # auto-detected based on first submission found
        lti_bearer_token = config_settings['lti_bearer_token']
        bearer_token_error_message = ('See the configuration file instructions, and the assignment\'s SpeedGrader '
                                      'page: %s/gradebook/speed_grader?assignment_id=%d') % (
                                         assignment_list_response_json[0]['html_url'].split('/assignments')[0],
                                         assignment_list_response_json[0]['id'])
        if lti_bearer_token.startswith('*** your'):
            print('WARNING: lti_bearer_token in', Config.FILE_PATH, 'seems to contain the example value.',
                  bearer_token_error_message)
        html_regex = re.compile('<.*?>')  # used to filter out HTML formatting from retrieved responses

        for quiz in assignment_list_response_json:
            quiz_id = quiz['id']
            print('\nFound new quiz with assignment ID', quiz_id, 'due at', quiz['due_at'])

            current_group = int(quiz['name'].split('[')[-1].rstrip(']').split(' ')[-1])
            valid_members = [g['student_number'] for g in groups[current_group]]
            print('\tIdentified group', current_group, 'with expected members', valid_members)

            assignment_url = Utils.course_url_to_api(quiz['html_url'])
            print('\tRequesting new quiz assignment submissions list from', assignment_url)
            with open(os.devnull, 'w') as f, contextlib.redirect_stdout(f):
                submission_list_response = Utils.get_assignment_submissions(assignment_url)
            if not submission_list_response:
                print('\tERROR: unable to retrieve new quiz assignment submission list')
                sys.exit()

            submission_list_json = json.loads(submission_list_response)
            user_session_map = []
            for submission_summary in submission_list_json:
                if submission_summary['submission_type'] and 'external_tool_url' in submission_summary:
                    current_rater = submission_summary['user']['login_id']
                    current_rater_name = submission_summary['user']['name']
                    if current_rater not in expected_submissions:
                        print('\tWARNING: skipping unexpected new quiz from student not in any group:', current_rater)
                        invalid.append(current_rater)
                        continue
                    if current_rater not in valid_members:
                        print('\tWARNING: skipping unexpected new quiz from student not in current group:',
                              current_rater)
                        invalid.append(current_rater)
                        continue
                    if submission_summary['workflow_state'] not in ['complete', 'graded', 'pending_review']:
                        print('\tWARNING: skipping empty or partly-complete new quiz from', current_rater, '-',
                              submission_summary)
                        invalid.append(current_rater)
                        continue

                    # date is oddly sometimes missing even if previously set
                    due_date = quiz['due_at'] or submission_summary['cached_due_date']
                    if due_date:
                        if (datetime.datetime.strptime(submission_summary['submitted_at'], TIMESTAMP_FORMAT) >
                                datetime.datetime.strptime(due_date, TIMESTAMP_FORMAT)):
                            print('\tWARNING: skipping late new quiz submission from', current_rater,
                                  '- submitted at',
                                  submission_summary['submitted_at'], 'but due at', due_date)
                            invalid.append(current_rater)
                            continue
                    print('\tFound new quiz submission from', current_rater_name, '- Canvas ID:',
                          submission_summary['user_id'], '; student number:', current_rater)

                    tool_url = submission_summary['external_tool_url']
                    tool_url_parts = tool_url.split('.quiz-lti-dub-')
                    user_session_map.append({'student_number': current_rater,
                                             'canvas_id': submission_summary['user_id'],
                                             'session_id': tool_url.split('participant_session_id=')[1].split('&')[0]})
                    if not lti_institution_subdomain:
                        lti_institution_subdomain = tool_url_parts[0].split('//')[1]
                    if not lti_environment_type:
                        lti_environment_type = tool_url_parts[1].split('.instructure.com')[0]

            if len(user_session_map) <= 0:
                print('\tNo valid submissions found for new quiz', quiz_id, '- skipping')
                continue

            current_responses = []
            current_errors = []
            current_total = 0
            found_members = []
            invalid_response = False

            lti_api_root = root_instructure_domain % (lti_institution_subdomain, 'lti', lti_environment_type)
            quiz_api_root = root_instructure_domain % (lti_institution_subdomain, 'api', lti_environment_type)

            token_headers = requests.structures.CaseInsensitiveDict()
            token_headers['accept'] = 'application/json'
            token_headers['authorization'] = ('%s' if 'Bearer ' in lti_bearer_token else 'Bearer %s') % lti_bearer_token

            for session in user_session_map:
                print('\t\tLoading new quiz session', session)
                token_response = requests.get(
                    '%s/participant_sessions/%s/grade' % (lti_api_root, session['session_id']), headers=token_headers)
                if token_response.status_code != 200:
                    print('\t\tERROR: unable to load new quiz session - did you set a valid lti_bearer_token in',
                          '%s?' % Config.FILE_PATH, bearer_token_error_message)
                    sys.exit()

                # first we get a per-submission access token
                attempt_json = token_response.json()
                quiz_session_headers = requests.structures.CaseInsensitiveDict()
                quiz_session_headers['accept'] = 'application/json'
                quiz_session_headers['authorization'] = attempt_json['token']
                quiz_session_id = attempt_json['quiz_api_quiz_session_id']

                # then a summary of the submission session and assignment overview
                submission_response = requests.get('%s/quiz_sessions/%d/' % (quiz_api_root, quiz_session_id),
                                                   headers=quiz_session_headers)
                if submission_response.status_code != 200:
                    print('\t\tERROR: unable to load quiz metadata - aborting:', submission_response)
                    sys.exit()

                submission_summary_json = submission_response.json()
                results_id = submission_summary_json['authoritative_result']['id']
                current_rater = session['student_number']
                current_rater_name = submission_summary_json['metadata']['user_full_name']
                print('\t\tLoaded new quiz submission summary', quiz_session_id, 'from', current_rater_name,
                      '- Canvas ID:', session['canvas_id'], '; student number:', current_rater)

                # then all quiz questions
                question_student_map = {}
                comments_question_id = None
                quiz_questions_response = requests.get(
                    '%s/quiz_sessions/%d/session_items' % (quiz_api_root, quiz_session_id),
                    headers=quiz_session_headers)
                quiz_question_response_json = quiz_questions_response.json()
                print('\t\tFound', end=' ')
                for question in quiz_question_response_json:
                    question_id = question['item']['id']
                    question_name = question['item']['title']
                    rating_question = False
                    if question['item']['user_response_type'] == 'Uuid':
                        rating_question = True
                        question_student_map[question_id] = {'subject': question_name,
                                                             'choices': question['item']['interaction_data']['choices']}
                    else:
                        comments_question_id = question['item']['id']  # assume there will be only one comments question
                    print('%s question' % ('rating' if rating_question else 'comments'), question_id,
                          'titled:', question_name, end='; ')
                print()

                # then all submissions for that quiz
                quiz_answers_response = requests.get(
                    '%s/quiz_sessions/%d/results/%s/session_item_results' % (
                        quiz_api_root, quiz_session_id, results_id),
                    headers=quiz_session_headers)
                submission_answers = quiz_answers_response.json()

                for answer in submission_answers:
                    if answer['item_id'] in question_student_map:
                        response_choices = question_student_map[answer['item_id']]
                        rated_student = response_choices['subject']
                        found_members.append(rated_student)

                        # validate the submitted data against Canvas group membership (remembering this may change)
                        if rated_student not in valid_members:
                            print('\t\t\tWARNING: Ignoring rating by', current_rater, 'of non member', rated_student)
                            continue

                        # for multiple choice responses we have to cross-reference the list of choices available
                        score = None
                        for value in answer['scored_data']['value']:
                            if answer['scored_data']['value'][value]['user_responded']:
                                for choice in response_choices['choices']:
                                    if choice['id'] == value:
                                        try:
                                            score = int(re.sub(html_regex, '', choice['item_body'].split(':')[0]))
                                            print('\t\t\tRating from', current_rater_name, 'for',
                                                  response_choices['subject'], ':', score)

                                            current_responses.append(
                                                [current_rater, rated_student, score, None, current_group])
                                            current_total += score
                                            break
                                        except ValueError as e:
                                            print('ERROR: Unable to process new quiz rating', value, '- has this quiz',
                                                  'been edited outside of the WebPA script? Error message:', e)
                                            sys.exit()
                            if score:
                                break

                        if not score:
                            print('\t\t\tWARNING: Unable to find matching choice from', current_rater_name, 'for',
                                  rated_student, 'rating; skipping:', answer)
                            current_errors.append('Member %s rating missing' % rated_student)
                            invalid_response = True
                            continue

                    elif answer['item_id'] == comments_question_id:
                        raw_answer = answer['scored_data']['value']
                        if raw_answer:
                            answer_text = re.sub(html_regex, '', raw_answer)
                            if answer_text and answer_text.lower().strip() != 'none':
                                print('\t\t\tWARNING: Comments from', current_rater_name, ':', answer_text)

                # finally, check for errors and collate responses
                # noinspection DuplicatedCode
                if current_group:
                    sorted_found = sorted(found_members)
                    sorted_expected = sorted(valid_members)
                    if sorted_found != sorted_expected:
                        members_missing = set(sorted_expected) - set(sorted_found)
                        if members_missing:
                            current_errors.append('Group member(s) missing: %s' % ', '.join(members_missing))
                            invalid_response = True
                        members_added = set(sorted_found) - set(sorted_expected)
                        if members_added:  # note: this can have legitimate explanations - e.g., members withdrawing
                            current_errors.append(
                                'Non-group member(s) found: %s – ignoring' % ', '.join(str(m) for m in members_added))

                if current_errors:
                    errors[current_rater] = current_errors

                if not invalid_response:
                    respondents.append(current_rater)
                    if current_errors:
                        print('\tWARNING: form data required corrections', current_rater, '-', current_errors)
                    for response in current_responses:
                        response[3] = response[2] / current_total
                        summary_sheet.append(response)
                else:
                    print('\tERROR: skipping invalid form from', current_rater, '-', current_errors)
                    invalid.append(current_rater)

        return respondents, invalid, errors

    @staticmethod
    def delete_quizzes(quiz_group_name):
        assignment_group_id = GroupResponseProcessor.get_assignment_group_id(quiz_group_name)
        if not assignment_group_id:
            print('ERROR: unable to find quiz group name to delete:', quiz_group_name, '- aborting')
            sys.exit()

        assignment_list_response = Utils.canvas_multi_page_request(
            '%s/assignment_groups/%s/assignments' % (COURSE_URL, assignment_group_id), type_hint='assignment list')
        if not assignment_list_response:
            print('\tERROR: unable to get assignment list response for deletion; aborting')
            sys.exit()

        deletion_confirmed = False
        assignment_list_response_json = json.loads(assignment_list_response)
        for quiz in assignment_list_response_json:
            if 'quiz_id' in quiz:
                print('Found quiz ID', quiz['quiz_id'], '-', quiz['name'], 'with assignment ID', quiz['id'])
            elif quiz['is_quiz_lti_assignment']:
                print('Found new quiz', quiz['name'], 'with assignment ID', quiz['id'])
            else:
                print('Skipping non-quiz assignment', quiz)
                continue

            if not args.dry_run and not deletion_confirmed:
                print()
                # noinspection SpellCheckingInspection
                if input('Confirm deleting quiz "%s" and all others in course %s,\nassignment group "%s" '
                         '(type yes or no) ' % (quiz['name'], COURSE_URL, quiz_group_name)).lower() != 'yes':
                    sys.exit('ERROR: aborting deletion; confirmation refused')
                deletion_confirmed = True

            if args.dry_run:
                print('\tDRY RUN: skipping deletion of quiz', quiz)
                continue

            quiz_deletion_url = '%s/assignments/%d' % (COURSE_URL, quiz['id'])
            quiz_deletion_response = requests.delete(quiz_deletion_url, headers=Utils.canvas_api_headers())
            if quiz_deletion_response.status_code == 200:
                print('\tDeleted assignment at %s:' % quiz_deletion_url, quiz)
            else:
                print('\tWARNING: unable to delete assignment at %s:' % quiz_deletion_url,
                      quiz_deletion_response.text, '-', quiz)
        print('DRY RUN: would delete' if args.dry_run else 'Deleted', len(assignment_list_response_json),
              'quiz assignments in group', quiz_group_name)


args = Args.interactive(get_args)
COURSE_URL = Utils.course_url_to_api(args.group[0].split('/groups')[0])

if args.setup_quiz_delete_existing:
    GroupResponseProcessor.delete_quizzes(args.quiz_group_name)
    sys.exit()  # this option overrides all others

group_id, group_sets = Utils.get_course_groups(args.group[0])
if not group_id or not group_sets:
    print('ERROR: unable to get group set ID from given URL', args.group[0])
    sys.exit()
print('%s WebPA response forms for group set %s' % ('Creating' if args.setup else 'Processing', group_id))

TEMPLATE_FILE = args.setup_spreadsheet_template
working_directory = os.path.dirname(
    os.path.realpath(__file__)) if args.working_directory is None else args.working_directory
WORKING_DIRECTORY = os.path.join(working_directory, str(group_id))
if args.setup and args.setup == 'spreadsheet' and os.path.exists(WORKING_DIRECTORY):
    print('ERROR: WebPA setup output directory', WORKING_DIRECTORY, 'already exists - please remove or rename')
    sys.exit()
if not (args.setup and args.setup == 'quiz' and not args.setup_quiz_export_links):
    os.makedirs(WORKING_DIRECTORY, exist_ok=True)

# setup mode - generate empty templates, either personalised per student or general per group
if args.setup:
    if args.setup in ['quiz', 'newquiz']:
        # we need an assignment group to place the quizzes in (which we also use later for retrieval)
        assignment_group_name = args.quiz_group_name if args.quiz_group_name else '%s [%s]' % (
            WEBPA_QUIZ_GROUP, datetime.datetime.now(datetime.UTC).strftime(TIMESTAMP_FORMAT))
        assignment_group_identifier = GroupResponseProcessor.get_assignment_group_id(assignment_group_name)
        if assignment_group_identifier:
            print('Found existing assignment group:', assignment_group_name, 'with ID:', assignment_group_identifier)
        else:
            print('Existing assignment group not found; creating new group:', assignment_group_name)
            if args.dry_run:
                print('\tDRY RUN: skipping creation of new assignment group')
                assignment_group_identifier = -1
            else:
                assignment_group_identifier = GroupResponseProcessor.create_assignment_group(assignment_group_name)
                print('\tCreated new assignment group with ID', assignment_group_identifier)

        if args.setup == 'quiz':
            GroupResponseProcessor.setup_quizzes(group_sets, assignment_group_identifier)
        elif args.setup == 'newquiz':
            GroupResponseProcessor.setup_new_quizzes(group_sets, assignment_group_identifier)
    elif args.setup == 'spreadsheet':
        GroupResponseProcessor.setup_spreadsheets(group_sets)
    else:
        print('Error: no setup format specified; aborting')
    sys.exit()

# processing mode - first load the marks to use as the baseline
marks_map = {}
if args.marks_file:
    marks_file = os.path.join(WORKING_DIRECTORY, args.marks_file)
    marks_map = Utils.get_marks_mapping(marks_file)
    if marks_map:
        print('Loaded original marks mapping for', len(marks_map), 'submissions:', marks_map)
    else:
        print('ERROR: marks mapping file', args.marks_file, 'empty or not found in assignment directory at', marks_file,
              '- aborting')
        sys.exit()

# next, load responses and create a master spreadsheet containing all rater responses (for, e.g., manual verification)
response_summary_workbook = openpyxl.Workbook()
response_summary_sheet = response_summary_workbook.active
response_summary_sheet.title = 'WebPA response form summary'
response_summary_sheet.freeze_panes = 'A2'  # set the first row as a header
response_summary_sheet.append(['Rater', 'Subject', 'Rating', 'Normalised', 'Group'])

if args.quiz_group_name:
    respondent_list, skipped_respondents, submission_errors = (
        GroupResponseProcessor.get_quizzes(group_sets, response_summary_sheet, args.quiz_group_name))
else:
    respondent_list, skipped_respondents, submission_errors = (
        GroupResponseProcessor.get_spreadsheets(group_sets, response_summary_sheet))
if len(respondent_list) <= 0:
    print('\nERROR: unable to continue; no valid WebPA responses to analyse')
    sys.exit()

response_summary_file = os.path.join(WORKING_DIRECTORY, 'webpa-response-summary.xlsx')
response_summary_workbook.save(response_summary_file)
print('\nProcessed', len(respondent_list), 'valid submissions of', len(respondent_list) + len(skipped_respondents),
      'total;', 'combined responses saved to', response_summary_file)
print('Skipped', len(skipped_respondents), 'late, invalid or tampered submissions from:', skipped_respondents)

# finally, shape original marks according to the summary file of group member ratings (using pandas for ease)
data = pandas.read_excel(response_summary_file, dtype={'Rater': str, 'Subject': str})  # student number is a string

# 1) count unique group members and number of submissions to calculate an adjustment factor
unique_data = data.groupby('Group').nunique()
count_group_members = unique_data['Subject']
count_webpa_submissions = unique_data['Rater']
webpa_adjustment_factor = (count_group_members / count_webpa_submissions).to_frame('Adjustment')

# 2) add a column containing the sum of the (normalised) scores, weighted by the adjustment factor
response_data = data.groupby(['Group', 'Subject']).agg(Score=('Normalised', 'sum'))  # new column: Score
response_data['Score'] *= webpa_adjustment_factor['Adjustment']  # include response rate adjustment
response_data = response_data.join(webpa_adjustment_factor)
response_data = response_data.join(count_webpa_submissions.to_frame('Raters'))
response_data = response_data.join(count_group_members.to_frame('Members'))

# 3) add a column showing whether the subject themselves responded
respondent_summary = {}
for respondent in respondent_list:
    respondent_summary[respondent] = 'Y'
response_present = pandas.DataFrame.from_dict(respondent_summary, orient='index')
response_present.rename(columns={response_present.columns[0]: 'Responded'}, inplace=True)
response_present.index.names = ['Subject']
response_data = response_data.join(response_present)
response_data = response_data[response_data.columns.tolist()[::-1]]  # reverse the column order for better display

# 4) add a column containing the standard deviation of the group's scores
webpa_variance = response_data.groupby('Group').agg(Variance=('Score', 'std'))  # new column: Variance
response_data['Variance'] = 1  # multiplication correctly maps group numbers; assignment does not, hence we initialise
response_data['Variance'] *= webpa_variance['Variance']

# 5) add any missing students/groups, then import the original marks (either individual or group)
original_marks = pandas.DataFrame.from_dict(marks_map, orient='index')
original_marks.rename(columns={original_marks.columns[0]: 'Original'}, inplace=True)
for group in group_sets.keys():
    if group not in response_data.index:  # add an empty dataframe with only the index values (group and student number)
        members = [(group, num['student_number']) for num in group_sets[group]]
        empty_row = pandas.DataFrame([[numpy.nan] * len(response_data.columns)], columns=list(response_data),
                                     index=pandas.MultiIndex.from_tuples(members, names=['Group', 'Subject']))
        response_data = pandas.concat([response_data, empty_row])

if all([not mark_key.isdigit() for mark_key in original_marks.index]):  # detect groups (student numbers are digits)
    original_marks.index = original_marks.index.to_series().str.replace(r'.+?(\d+)', lambda m: m.group(1),
                                                                        regex=True).astype(int)
    original_marks.index.names = ['Group']
else:  # marks file is individual students
    original_marks.index.names = ['Subject']
original_marks = original_marks.filter(['Original'])  # keep only the index and the mark colum; drop all others
response_data = response_data.join(original_marks)
response_data = response_data.sort_values(['Group', 'Subject'])

# 6) if variance is above the threshold, adjust marks according to the weighting; otherwise use the original group mark
response_data['Weighted'] = response_data['Original']
response_data.loc[response_data['Variance'] >= args.minimum_variance, 'Weighted'] *= response_data['Score']

# 7) ensure marks do not go above the maximum for the assignment, round to the nearest 0.5 and highlight issues
response_data['Mark'] = response_data['Weighted']
response_data.loc[response_data['Mark'] > args.maximum_mark, 'Mark'] = args.maximum_mark
rounding_factor = 1 / args.mark_rounding  # e.g., 0.5 -> 2 to round to nearest 0.5
try:
    response_data['Mark'] = (response_data['Mark'] * rounding_factor).round().astype(int) / rounding_factor
except pandas.errors.IntCastingNaNError:
    print('ERROR: unable to round marks, probably due to a group name mismatch or missing mark spreadsheet row. Have',
          'you correctly named groups in the `--marks-file` provided? (Note that group names must *exactly* match the',
          'names used on Canvas)')
    raise
response_data['Scaled'] = response_data.apply(
    lambda x: 'Y' if not math.isclose(x['Original'], x['Weighted'], abs_tol=0.00001) else '', axis=1)
if args.context_summaries:
    response_data['Errors'] = pandas.NA
    response_data['Comment'] = pandas.NA

# 8) save to a calculation result file, highlighting errors, missing data and scaled values, and context if requested
output_file = os.path.join(WORKING_DIRECTORY, 'webpa-calculation.xlsx')
writer = pandas.ExcelWriter(output_file, engine='openpyxl')
response_data.to_excel(writer, sheet_name='WebPA calculation')

for row in writer.book.active.iter_rows(min_row=2, min_col=3, max_col=3):
    if not row[0].value:
        row[0].fill = openpyxl.styles.PatternFill(start_color='00FFC7CE', end_color='00FFC7CE', fill_type='solid')
for row in writer.book.active.iter_rows(min_row=2, min_col=2, max_col=12):
    if row[10].value == 'Y':
        row[10].fill = openpyxl.styles.PatternFill(start_color='00FFB97F', end_color='00FFB97F', fill_type='solid')
        for respondent in skipped_respondents:
            if row[0].value == respondent:
                print('WARNING: Respondent who submitted an invalid form had their mark adjusted:', row[0].value)
                break

if args.context_summaries:
    for row in writer.book.active.iter_rows(min_row=2, min_col=2, max_col=14):
        for key in submission_errors.keys():
            if row[0].value == key:
                row[11].value = '; '.join(submission_errors[key])  # doesn't seem to be an easy way to do with pandas
                break
        for respondent in respondent_list:
            if row[0].value == respondent:
                if row[11].value:
                    row[12].value = 'You submitted a valid contribution form that required correction for the ' \
                                    'following reason(s): %s.' % row[11].value
                else:
                    row[12].value = 'You submitted a valid contribution form.'
                break
        if not row[12].value:
            if row[11].value:
                row[12].value = 'You submitted a contribution form, but it was invalid for the following reason(s): ' \
                                '%s.' % row[11].value
            else:
                row[12].value = 'You did not submit a contribution form.'

writer.close()

print('\nSuccessfully calculated WebPA scores and saved calculation to', output_file, '- summary:')
print(response_data)

# because we add comments using openpyxl, we need to reopen the workbook to save the final version with comments
scaled_marks_file = os.path.join(WORKING_DIRECTORY, 'webpa-final-marks.xlsx')
scaled_marks_title = 'WebPA results'
if args.context_summaries:
    result_summary_workbook = openpyxl.load_workbook(output_file)
    result_summary_sheet = result_summary_workbook[result_summary_workbook.sheetnames[0]]
    result_summary_sheet.title = scaled_marks_title
    for merge in list(result_summary_sheet.merged_cells):  # need to unmerge or subject column inherits group merge
        result_summary_sheet.unmerge_cells(range_string=str(merge))
    result_summary_sheet.delete_cols(12, 2)  # calculation comments (remove in reverse to preserve index numbers)
    result_summary_sheet.delete_cols(3, 8)  # calculation details
    result_summary_sheet.delete_cols(1, 1)  # group number
    result_summary_workbook.save(scaled_marks_file)
else:
    result_summary = response_data.filter(['Subject', 'Mark'], axis=1)
    result_summary.to_excel(scaled_marks_file, sheet_name=scaled_marks_title)
print('Saved WebPA-adjusted marks to', scaled_marks_file)
